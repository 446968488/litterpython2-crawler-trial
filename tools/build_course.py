# -*- coding: utf-8 -*-
"""爬虫实战课程生成器（14 岁以上 · 幽默梗王风 · 纯图文无视频）。

用 Python 写内容、json 序列化输出，彻底避开 JS 字符串转义坑。
运行: python tools/build_course.py
产物: data/course.js, data/talk.js, data/audio.js, data/words.js, data/vocab-words.js

约定:
- markdown 一律用 r\"\"\" 原始串，避免 \\n \\d 等被吞。
- 题目类型: choice / fill / open / order / typing / coding / tap
  * choice: {type,question,options,answer(下标),explain}
  * fill:   {type,question,answer(支持 | 多解, 小写精确匹配),explain}
  * open:   {type,question,answer(参考答案文本)}
  * order:  {type,question,steps:[...],explain}
  * typing: {type,question,words:[大写英文]}
  * coding: {type,question,starter,expect(输出子串),hint}
  * tap:    {type,question,options,multi,answer:[下标...],explain}
- 字段: id,title,icon,markdown,figures,takeaway,words,exercises,code,tasks
  * 本套课【不设置 video】，纯图文。
  * code 字段 = 页面内"真跑代码"沙盒(走 Skulpt，只能纯 Python)。
"""

import json
import os

try:
    from eng_to_ipa import convert as _ipa_convert
    def _to_ipa(word):
        try:
            return _ipa_convert(word)
        except Exception:
            return ''
except ImportError:
    def _to_ipa(word):
        return ''

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'data')

CHAPTER_COLOR = {
    '入门启航': '#5b8fc4',
    'Python 速通': '#3a9d5d',
    '网页与 HTTP': '#e0922f',
    'requests 抓取': '#d9536b',
    'BeautifulSoup 解析': '#7a5fb0',
    '正则与 JSON': '#2f9e9e',
    '数据存储': '#c77dff',
    '进阶技巧': '#ef6f53',
    '反爬与合规': '#6c8ae4',
    '实战项目': '#e0567a',
    '毕业冲刺': '#39a0a8',
}


def L(id, title, icon, md, *, figures=None, takeaway, words=None,
      exercises=None, code=None, tasks=None, color=None):
    d = {'id': id, 'title': title, 'icon': icon, 'markdown': md,
         'takeaway': takeaway}
    if figures:
        d['figures'] = figures
    if words:
        d['words'] = words
    if exercises:
        d['exercises'] = exercises
    if code is not None:
        d['code'] = code
    if tasks:
        d['tasks'] = tasks
    if color:
        d['color'] = color
    return d


def CH(title, lessons):
    return {'title': title, 'lessons': lessons}


# ============================ 第0章 入门启航 ============================
ch0 = CH('入门启航', [
    L('r0l1', '爬虫是什么：自动化的你', '🤖',
      r"""## 爬虫，就是"不知疲倦的你"

你刷网页的过程，本质是一串机械动作：**输入网址 → 浏览器发请求 → 服务器回 HTML → 渲染成页面 → 你用眼睛抠信息**。

爬虫干的事一模一样，只是把"你用眼睛抠"换成"代码自动抠"，而且能**重复成千上万次、从不喊累、不摸鱼**。

### 一句话定义
> 爬虫 = 会发 HTTP 请求 + 会读 HTML/JSON + 会把想要的信息抠出来存好的程序。

### 它和"搜索引擎"是亲戚
百度、谷歌每天派无数爬虫把全网扫一遍建索引——你写的爬虫，就是**私人的、只为你服务的迷你搜索引擎**。

### 别想太玄
很多人一听"爬虫"就觉得是黑客技术。真不是。它就是**请求、解析、存储**三步循环，配上一点点耐心和礼貌。后面几十节，就是把这个循环练到肌肉记忆。

| 你手动做 | 爬虫帮你做 |
|---|---|
| 打开浏览器、点网址 | 代码 `requests.get(url)` |
| 肉眼找价格/标题 | 代码按规则提取 |
| 手抄到 Excel | 代码写 CSV/JSON |""",
      figures=[{'key': 'browser_server', 'caption': '🌐 你点网址→浏览器发请求→服务器回HTML→渲染；爬虫用代码做同样的事，跳过人眼直接抠数据'}],
      takeaway=r"""爬虫没啥神秘的：**发请求、收 HTML、抠数据、存起来**，循环而已。它就是不知疲倦的你，不是黑客魔法。把"请求-解析-存储"这六个字刻进脑子，后面全是在给它加细节。""",
      words=[
          {'en': 'CRAWLER', 'zh': '爬虫：自动抓取网页数据的程序，也叫 spider（蜘蛛）', 'pron': 'ˈkrɔːlər'},
          {'en': 'REQUEST', 'zh': '请求：爬虫发给服务器的"我要这个页面"的消息', 'pron': 'rɪˈkwest'},
          {'en': 'PARSE', 'zh': '解析：从原始文本里把想要的结构化信息抠出来', 'pron': 'pɑːrs'},
      ],
      exercises=[
          {'type': 'choice', 'question': '爬虫和真人刷网页最大的区别是？', 'options': ['爬虫跳过了"人眼看"，直接拿 HTML 抠数据', '爬虫必须用 IE 浏览器', '爬虫不联网'], 'answer': 0, 'explain': '爬虫省掉了人工浏览，直接处理服务器返回的 HTML 文本。'},
          {'type': 'choice', 'question': '下面哪句话对爬虫描述是错的？', 'options': ['爬虫能重复成千上万次不累', '爬虫本质也是发请求收响应', '爬虫一定是非法的黑客工具'], 'answer': 2, 'explain': '爬虫本身中立，合不合法取决于你抓什么、怎么抓、用不用礼貌。'},
          {'type': 'tap', 'question': '爬虫的标准三步循环包含哪些？（多选）', 'options': ['发请求', '解析提取', '存数据', '雇人抄写'], 'multi': True, 'answer': [0, 1, 2], 'explain': '标准循环就是 请求 → 解析 → 存储；"雇人抄写"是没写爬虫时的原始人操作。'},
          {'type': 'fill', 'question': '爬虫能帮个人做的"迷你搜索引擎"，本质是只为你服务的______（填两个字：爬/搜/存中的一个角色）。', 'answer': '爬虫', 'explain': '爬虫就是私人的、只为你服务的迷你搜索引擎。'},
          {'type': 'open', 'question': '你觉得"写爬虫"最可能被误会的点是什么？用自己的话写一句（比如别人以为你是黑客）。', 'answer': '爬虫常被误以为是黑客技术，其实只是自动发请求、解析、存数据的循环程序，本身中立。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['CRAWLER', 'REQUEST', 'PARSE']},
      ],
      tasks=[
          '打开浏览器，随便进一个你常逛的网站，按 F12 打开开发者工具，切到 Network 面板刷新一次，看它发了多少个请求——这就是爬虫要模拟的动作。',
          '在本子上画一张"请求→响应→解析→存储"的草图，贴显示器边上，整个课程都围绕它。',
          '想一个你真正想抓的数据（天气？某站榜单？表情包？），写一句话目标，后面每学一节就往这个目标靠近一点。',
      ]),

    L('r0l2', '学完你能薅到什么：典型场景', '🎯',
      r"""## 学完这套课，你能整这些活

别被"爬虫"俩字吓到，它落地的场景其实很生活化。下面这些是**学完就能做**的：

### 🌤️ 1. 个人数据看板
自动抓天气、汇率、股价、快递，攒成自己的小仪表盘，每天打开就是最新数据，不用手动刷。

### 📚 2. 榜单 / 列表 → Excel
豆瓣电影 Top250、某电商商品列表、公众号文章清单……一键存成 `.csv` / `.xlsx`，Excel 直接打开分析。

### 🖼️ 3. 批量下载
一整套头像、表情包、壁纸、教材插图，循环几行代码全部搬回家，告别手动一张张另存为。

### 🔎 4. 从接口里"白嫖"结构化数据
很多 App / 网站的数据其实是通过接口（API）返回的 JSON。直接打接口，比解析 HTML 还省事——这是进阶最爱。

### ⚠️ 一条铁律先说在前
**能抓 ≠ 该抓**。个人隐私、付费内容、人家明确禁止的，碰都别碰。合规那一章（第 8 章）会讲清楚红线，先有个"不是啥都能爬"的意识。

> 这一节没有代码，先把"我想抓什么"想明白。目标越具体，后面学得越带劲。""",
      figures=[{'key': 'crawler_data_flow', 'caption': '🕸️ 爬虫全流程：发出请求 → 拿到网页/接口 → 解析提取 → 存成文件/数据库，闭环'}],
      takeaway=r"""学完你能干四件实在事：做个人数据看板、把榜单存 Excel、批量下载图片、直接薅接口 JSON。**但记住：能抓不等于该抓**，隐私和人家禁抓的别碰，第 8 章专门讲红线。""",
      words=[
          {'en': 'API', 'zh': '接口：网站用来拉取数据的网址，常返回现成的 JSON'},
          {'en': 'DASHBOARD', 'zh': '看板：把多个数据集中展示的页面/面板'},
          {'en': 'BATCH', 'zh': '批量：一次性对一堆目标做同样操作'},
      ],
      exercises=[
          {'type': 'tap', 'question': '以下哪些是学完本套课能做的？（多选）', 'options': ['做天气/汇率个人看板', '把电影榜单存成 Excel', '批量下载图片', '黑进银行系统'], 'multi': True, 'answer': [0, 1, 2], 'explain': '前三个都是正经爬虫落地场景；第四个是犯罪，不在课程范围也不在地球范围。'},
          {'type': 'choice', 'question': '从网站接口（API）直接拿数据，相比解析 HTML 通常？', 'options': ['更省事，数据已经是结构化 JSON', '更麻烦', '完全不可能'], 'answer': 0, 'explain': '接口返回的 JSON 本身就是结构化数据，省去解析 HTML 的麻烦。'},
          {'type': 'choice', 'question': '课程反复强调的"能抓 ≠ 该抓"是在提醒你注意？', 'options': ['合规与法律红线', '电脑性能', '网速'], 'answer': 0, 'explain': '爬虫能力再强，也要守住隐私、付费内容、禁抓声明这些红线。'},
          {'type': 'open', 'question': '你最想用爬虫解决自己的哪个具体问题？写一两句（越具体越好，比如"每周一把关注的 5 个 UP 主更新汇总成表"）。', 'answer': '目标示例：每周一把关注的 UP 主更新汇总成表 / 自动抓天气发到群里 / 把喜欢的图集批量下载。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['API', 'DASHBOARD', 'BATCH']},
      ],
      tasks=[
          '把"我最想抓的一个目标"写成一句话需求文档（含：数据来源网站、想要哪些字段、最终存成什么）。',
          '在开发者工具 Network 里找一个返回 JSON 的接口（很多网站列表页都有），把它的 URL 抄下来备着。',
          '列 3 个"虽然想抓但可能不合适"的数据源（比如别人付费课程、私密相册），提醒自己避开。',
      ]),

    L('r0l3', '爬虫的规矩：法律·隐私·robots', '⚖️',
      r"""## 先立规矩，再写代码

写爬虫最容易翻车的不是技术，是**没数**。这一节不教你写代码，教你**别作死**。

### 🚫 绝对别碰的红线
- **个人信息**：身份证号、手机号、住址、聊天记录——抓这些可能直接触犯《个人信息保护法》，不是"平台封你"那么简单。
- **付费 / 登录墙内容**：人家卖钱的课程、会员视频，你爬了等于盗窃。
- **国家机密、色情、赌博等违法内容**：碰了就不是封号的事了。

### 🤝 基本的"爬虫礼仪"
- **看 `robots.txt`**：网站根目录下 `https://站点/robots.txt` 写明了"哪些能爬、哪些不能"。`Disallow` 的就是人家说不要。先读它。
- **别把人家服务器薅秃**：加延迟、控制并发，别一秒打几百次请求把人家打挂。
- **带上真实 UA、别伪装成浏览器骗人**：诚实一点，说明你是谁（后面讲）。
- **遵守版权**：抓来的内容别商用乱发。

### 一个心态
> 好的爬虫，是**有礼貌的访客**；坏的爬虫，是**半夜砸人家店门还顺走东西的混蛋**。咱当前者。

第 8 章会把 robots、速率限制、法律红线拆得更细，这一节先把"红线意识和礼貌意识"种进脑子里。""",
      figures=[
          {'key': 'legal_redline', 'caption': '🚫 红线：个人信息/付费内容/违法内容千万别碰；礼貌：看 robots、控速率、带真实 UA'},
          {'key': 'robots_txt', 'caption': '📄 robots.txt 是网站根的"门牌告示"：Disallow 的路径就是人家说"别爬这里"'},
      ],
      takeaway=r"""写爬虫先立规矩：**个人信息、付费内容、违法内容三条红线碰都别碰**；礼貌上先看 robots.txt、控速率别薅秃服务器、带真实 UA。好爬虫是有礼貌的访客，不是半夜砸店门的混蛋。""",
      words=[
          {'en': 'ROBOTS', 'zh': 'robots.txt：网站根的告示文件，写明哪些路径允许/禁止爬取'},
          {'en': 'PRIVACY', 'zh': '隐私：个人身份相关数据，抓取它法律风险极高'},
          {'en': 'COMPLIANCE', 'zh': '合规：遵守法律与网站规则，爬虫的长期生存之道'},
      ],
      exercises=[
          {'type': 'tap', 'question': '以下哪些属于"绝对别碰"的爬虫红线？（多选）', 'options': ['抓取他人手机号/身份证', '爬取付费会员视频', '爬取公开的新闻标题', '抓取色情赌博内容'], 'multi': True, 'answer': [0, 1, 3], 'explain': '个人信息、付费内容、违法内容都是红线；公开新闻标题一般没问题（但仍要看 robots 与版权）。'},
          {'type': 'choice', 'question': '网站根目录的 robots.txt 主要作用是？', 'options': ['告诉爬虫哪些路径允许/禁止爬取', '加快网页加载', '存储用户密码'], 'answer': 0, 'explain': 'robots.txt 是站长的"门牌告示"，Disallow 即"别爬这里"。'},
          {'type': 'choice', 'question': '为了不把人家服务器薅挂，正确做法是？', 'options': ['加请求延迟、控制并发', '一秒打几百次请求', '伪造一堆 UA 假装很多人'], 'answer': 0, 'explain': '礼貌爬虫要控速率；疯狂请求会打挂服务器，还可能被封甚至担责。'},
          {'type': 'fill', 'question': '网站用来声明"哪些路径禁止爬取"的文件叫 ______.txt。', 'answer': 'robots', 'explain': 'robots.txt 是标准约定，放在网站根目录。'},
          {'type': 'open', 'question': '有人说"反正网站是公开的，我想爬就能爬"。你怎么反驳他？写 2-3 句。', 'answer': '公开不等于可随意抓取：还要看 robots.txt、版权与是否合理（频率/量级），个人信息与付费内容即使"能看到"也受法律保护，抓取即违法。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['ROBOTS', 'PRIVACY', 'COMPLIANCE']},
      ],
      tasks=[
          '打开 `https://www.baidu.com/robots.txt`（或你常逛的某个站点根目录 /robots.txt），读一遍，找一条 Disallow 规则记下来。',
          '写下你的"爬虫行为准则"三条（例如：先看 robots、每秒不超过 N 次、绝不抓个人信息）。',
          '找一个你目标站点的 robots.txt，判断你的目标数据是否在禁止之列；若在，换个数据源或放弃。',
      ]),
])

# ============================ 第1章 Python 速通 ============================
ch1 = CH('Python 速通', [
    L('r1l1', '字符串与列表：处理文本的左手右手', '🔤',
      r"""## 爬虫天天和"文本"打交道

网页是文本，接口返回是文本，抠出来的标题/价格还是文本。所以**字符串**和**列表**是你最先要练熟的两样。

### 字符串：一串字符
```python
s = 'Python 爬虫很有趣'
print(s[0])        # P
print(s[0:6])      # Python（切片：从下标0取到6之前）
print(s.replace('有趣', '上头'))   # 替换
print('爬虫' in s) # True（是否包含）
```
> 切片 `s[a:b]` 是"从 a 取到 b 之前"，左闭右开。这是抠子串的万能钥匙。

### 列表：有序的一排数据
```python
titles = ['第一章', '第二章', '第三章']
titles.append('第四章')      # 加一个
print(len(titles))          # 4
for t in titles:            # 挨个处理
    print(t)
```

### 为什么重要
抓回来的书名、链接、价格，十有八九先装进**列表**，再统一存文件。上面那个框能真跑，改改试试。""",
      takeaway=r"""字符串会切片(`s[0:6]`)、会替换、会判断包含；列表会加元素、会循环。**爬虫抓回来的东西几乎都先切成字符串、再装进列表**——这俩是你的左手右手，练熟了后面解析数据顺得飞起。""",
      words=[
          {'en': 'STRING', 'zh': '字符串：用引号括起来的一串文字，如 "hello"'},
          {'en': 'LIST', 'zh': '列表：用方括号装的一排数据，可随时增减'},
          {'en': 'SLICE', 'zh': '切片：从序列切出一段，如 s[0:6] 左闭右开'},
      ],
      code=r"""titles = ['第一章', '第二章', '第三章']
titles.append('第四章')
print('共', len(titles), '章')
for t in titles:
    print(t)
s = 'Python 爬虫很有趣'
print(s[0:6])
print('爬虫' in s)
print(s.replace('有趣', '上头'))""",
      exercises=[
          {'type': 'choice', 'question': '想从字符串 s 里取出前 3 个字符，正确的是？', 'options': ['s[0:3]', 's[3]', 's[:3:1]'], 'answer': 0, 'explain': '切片 s[0:3] 表示"从下标0取到3之前"，正好3个字符。'},
          {'type': 'fill', 'question': '切片左闭右开，s[2:5] 实际取到的是下标 ____ 到 ____ 之前（填两个数字，用空格隔开如 "0 3"）。', 'answer': '2 5', 'explain': 's[a:b] 取 a 到 b 之前，即下标 2、3、4 三个字符。'},
          {'type': 'coding', 'question': '把下面三个城市名放进列表并打印数量，目标输出里出现「3」。', 'starter': "cities = ['北京', '上海', '广州']\n# 在这里打印城市数量（用 len）", 'expect': '3', 'hint': '用 print(len(cities)) 输出列表长度。'},
          {'type': 'tap', 'question': '关于 Python 列表，下面哪些说法对？（多选）', 'options': ['用方括号 [] 定义', 'append 能在末尾加元素', '长度固定不能改', '可以用 for 挨个遍历'], 'multi': True, 'answer': [0, 1, 3], 'explain': '列表可变、可 append、可遍历；"长度固定"是元组不是列表。'},
          {'type': 'open', 'question': '爬虫抓回 100 个商品价格，你打算先用什么结构装、再怎么存？写 2 句思路。', 'answer': '先用列表装下所有价格（或装"每条记录字典"的列表），再循环写进 CSV/JSON 文件落地。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['STRING', 'LIST', 'SLICE']},
      ],
      tasks=[
          '在上面的代码框里把替换的目标词改成别的（比如把“上头”换成“离谱”），再跑一次看效果。',
          '自己写 3 行：定义一个你喜欢的 5 部电影名列表，打印第 2 部，再 append 一部。',
          '找一个网页标题（随便复制一段文字），用切片把它前 10 个字截出来打印。',
      ]),

    L('r1l2', '字典：给数据贴标签', '📒',
      r"""## 一条数据，天然是"名字=值"

一本书有：书名、作者、价格。用**字典（dict）**最合适——**用键找值**，不用记位置。

```python
book = {'title': 'Python 入门', 'author': '小光', 'price': 39}
print(book['title'])     # Python 入门
book['price'] = 45       # 改值
book['stock'] = 10       # 加新键
print(book.keys())       # 所有键
```

### 爬虫为什么离不开字典
抓回来的每条记录，天然就是"字段名: 值"。把每条存成字典，再把这些字典装进列表，就是一份干净的数据表：

```python
books = [
    {'title': 'A', 'price': 30},
    {'title': 'B', 'price': 25},
]
for b in books:
    print(b['title'], b['price'])
```

后面存 CSV、JSON，本质都是在序列化这种**"字典的列表"**。记住这个形状，它能陪你走到毕业。""",
      takeaway=r"""字典用键存取，特别适合表示"一条记录的多个字段"；**字典的列表就是一张表**，是爬虫数据的标准形状。后面 CSV、JSON 本质都在序列化它。""",
      words=[
          {'en': 'DICT', 'zh': '字典：用 键:值 成对存储，按名字快速取，如 {"name":"小明"}'},
          {'en': 'KEY', 'zh': '键：字典里用来找值的名字'},
          {'en': 'VALUE', 'zh': '值：键对应的内容'},
      ],
      code=r"""book = {'title': 'Python 入门', 'author': '小光', 'price': 39}
book['price'] = 45
book['stock'] = 10
print(book['title'], book['price'])
books = [{'title': 'A', 'price': 30}, {'title': 'B', 'price': 25}]
for b in books:
    print(b['title'], b['price'])""",
      exercises=[
          {'type': 'choice', 'question': '想取出字典 d 里键为 "name" 的值，正确的是？', 'options': ['d.name', 'd["name"]', 'd(name)'], 'answer': 1, 'explain': '字典用方括号加键名取值：d["name"]。'},
          {'type': 'fill', 'question': '往字典 d 里新增一个键 age=14，应写 d[___] = 14。', 'answer': '"age"', 'explain': '用 d["age"]=14 新增键值对；引号不能省。'},
          {'type': 'coding', 'question': '创建一个字典表示一个人{name:小明, age:14}，并打印他的年龄，目标输出含「14」。', 'starter': "p = {'name': '小明', 'age': 14}\n# 打印 p 的 age", 'expect': '14', 'hint': 'print(p["age"]) 即可。'},
          {'type': 'tap', 'question': '关于"字典的列表"这种数据结构，说法对的有？（多选）', 'options': ['每个字典代表一条记录', '整体是一个列表=一张表', '只能存数字', '适合直接存成 JSON'], 'multi': True, 'answer': [0, 1, 3], 'explain': '字典=一条记录，列表=多行，JSON 原生支持这种嵌套结构。'},
          {'type': 'open', 'question': '为什么爬虫抓"书单"时，用 [{书名,作者,价格}, ...] 比用平行的三个列表更好？写 2 句。', 'answer': '字典把一条记录的字段绑在一起不会错位；整体列表就是一张表，直接序列化 JSON/CSV，比三个平行列表更不易出错。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['DICT', 'KEY', 'VALUE']},
      ],
      tasks=[
          '在代码框里给 books 再加一本 {title:"C", price:40}，打印每本的平均价。',
          '把自己手机里的 3 个 App 写成 [{name, 用途}, ...] 的字典列表。',
          '想一个"一条记录多个字段"的真实场景（如购物车商品），用字典列表描述它。',
      ]),

    L('r1l3', '文件读写与异常：别让脏数据崩了你', '📂',
      r"""## 爬虫最后都要"落地"成文件

存数据、读缓存，都离不开文件操作。同时网络会超时、网页会缺字段——用**异常**兜住，程序才稳。

### 写文件（中文必加 utf-8）
```python
with open('data.txt', 'w', encoding='utf-8') as f:
    f.write('你好，爬虫\n')
```
`with` 自动关文件，`encoding='utf-8'` 防中文乱码——**爬虫写中文永远带它**。

### 读文件
```python
with open('data.txt', 'r', encoding='utf-8') as f:
    text = f.read()
```

### 异常：预料之中的意外
```python
try:
    price = int('暂无')     # 会炸
except ValueError:
    price = 0              # 兜住，继续跑
```
抓网页时字段缺失、转数字失败太常见了。`try/except` 让程序跳过坏数据继续，而不是**一条脏数据全线崩溃**。

### 小提醒
上面 `code` 框跑的是"内存里模拟脏数据"的版本（浏览器里的 Python 不能真写你硬盘），但语法和实际一模一样，回到自己电脑上把 `open` 打开就能真存。""",
      takeaway=r"""`with open(..., encoding='utf-8')` 读写文件防乱码；`try/except` 兜住解析失败。**爬虫遇到脏数据别崩，跳过继续**——一条坏数据不该拖垮全场。""",
      words=[
          {'en': 'FILE', 'zh': '文件：把数据存到磁盘的载体，如 data.txt'},
          {'en': 'UTF8', 'zh': 'UTF-8：最常用字符编码，写中文务必指定它防乱码'},
          {'en': 'EXCEPT', 'zh': '异常捕获：用 try/except 接住可能出错的地方'},
      ],
      code=r"""text = '价格:39\n价格:25\n价格:暂无'
lines = text.strip().split('\n')
total = 0
for line in lines:
    num = line.split(':')[1]
    try:
        total += int(num)
    except ValueError:
        print('跳过一条脏数据:', num)
print('合计', total)""",
      exercises=[
          {'type': 'choice', 'question': '写中文文件时，下面哪项最该加？', 'options': ['encoding="utf-8"', 'mode="big"', 'speed="fast"'], 'answer': 0, 'explain': '指定 encoding="utf-8" 才能正确存中文，否则容易乱码。'},
          {'type': 'fill', 'question': '要把可能出错的 int(x) 包起来，应该用 ___ / except 结构。', 'answer': 'try', 'explain': 'try/except 用来捕获和处理异常。'},
          {'type': 'tap', 'question': '关于异常捕获，说法对的有？（多选）', 'options': ['try 里放可能出错的代码', 'except 里写兜底处理', '有了它程序就永远不报错', '能防止一条脏数据崩全场'], 'multi': True, 'answer': [0, 1, 3], 'explain': 'try/except 接住异常让程序继续；但"永远不报错"是错觉，该崩的逻辑错误仍会暴露。'},
          {'type': 'open', 'question': '你抓的一页里，某商品价格写的是"促销价"而不是数字，你的代码该怎么处理才不崩？写 2-3 行思路。', 'answer': '用 try/except 包住 int(价格)，转换失败时给默认值 0 或跳过该条，并打印日志，保证其他数据继续处理。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['FILE', 'UTF8', 'EXCEPT']},
      ],
      tasks=[
          '在自己电脑上新建 test.txt，用 `with open` 写三行中文，再读回来打印（注意带 utf-8）。',
          '故意把一段"价格:暂无"喂给 int()，用 try/except 接住并打印"跳过"。',
          '想一个你爬虫里可能遇到的脏数据场景（缺失字段/格式错），写下兜底方案。',
      ]),

    L('r1l4', '函数、模块与 pip：装上 requests 这把枪', '📦',
      r"""## 把动作打包成函数，把工具装在模块里

### 函数：一次定义，到处调用
```python
def clean(text):
    return text.strip().replace('\n', ' ')

print(clean('  hello  '))
```
函数让你把"清洗文本"这种重复动作收成一个名字，后面调用一行搞定。

### 模块：别人写好的工具箱
Python 自带 `re`（正则）、`json`、`csv`。**第三方库**用 `pip` 安装：
```bash
pip install requests beautifulsoup4 lxml
```
装好后 `import` 即用：
```python
import requests
from bs4 import BeautifulSoup
```

### 为什么这节没代码框
`requests`/`bs4` 要连真实网络、装了第三方库，浏览器里那个迷你 Python 跑不了。所以这一节是**让你回自己电脑把枪装上**——后面所有实战都靠它们。

> 安装慢？加国内镜像：`pip install requests -i https://pypi.tuna.tsinghua.edu.cn/simple`""",
      takeaway=r"""函数把重复动作打包；模块是别人写好的工具箱；**requests、BeautifulSoup 用 pip 装好才能用**——这一下是整个爬虫世界的入场券。装完再往下学。""",
      words=[
          {'en': 'FUNCTION', 'zh': '函数：用 def 定义、可重复调用的代码块'},
          {'en': 'MODULE', 'zh': '模块：一个 .py 工具箱，import 后使用'},
          {'en': 'PIP', 'zh': 'pip：Python 的包管理器，用来安装第三方库'},
      ],
      exercises=[
          {'type': 'choice', 'question': '要安装第三方库 requests，正确命令是？', 'options': ['pip install requests', 'import requests', 'open requests'], 'answer': 0, 'explain': 'pip install 负责安装，import 负责使用，两步走。'},
          {'type': 'fill', 'question': '用 def 定义一个名为 add 的函数，第一行应写 ___ add(a, b):', 'answer': 'def', 'explain': '函数用 def 关键字定义。'},
          {'type': 'tap', 'question': '关于 pip 和模块，说法对的有？（多选）', 'options': ['pip 用来装第三方库', '装完用 import 使用', 're/json 是自带模块', '函数必须联网才能用'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'pip 装第三方、import 使用、re/json 自带；函数纯本地，不联网。'},
          {'type': 'open', 'question': '你在自己电脑上 pip install 报错"找不到命令"，最可能的原因和解决办法是什么？写 2 句。', 'answer': '多半是 pip 没进 PATH 或用了多个 Python；可试 python -m pip install 包名，或确认 pip 对应正确版本的 Python。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['FUNCTION', 'MODULE', 'PIP']},
      ],
      tasks=[
          '在自己电脑终端执行 `pip install requests beautifulsoup4 lxml`，确认无报错（可用 `python -m pip` 兜底）。',
          '开 Python 敲 `import requests; print(requests.__version__)`，能打印版本号即装好。',
          '写一个 clean(text) 函数，把一段带多余空格和换行的文本压成单行，本地跑通。',
      ]),

    L('r1l5', '推导式与 lambda：一行顶十行的偷懒艺术', '🪄',
      r"""## 写爬虫要学会"偷懒"

同样的结果，能用一行写完就别写五行。两个神器：**列表推导式** 和 **lambda**。

### 列表推导式：批量加工的语法糖
```python
prices = ['39', '25', '暂无']
nums = [int(p) for p in prices if p.isdigit()]   # [39, 25]
```
一行把"字符串列表"过滤+转数字，干净利落。爬虫里抠完一堆文本后常这么收拾。

### lambda：临时小函数
```python
books = [{'t': 'B', 'p': 25}, {'t': 'A', 'p': 30}]
books.sort(key=lambda x: x['p'])     # 按价格排序
print([b['t'] for b in books])       # ['B', 'A']
```
`lambda x: x['p']` 是个"用完即弃"的小函数，常当排序/筛选的钥匙。

### 别滥用
推导式套三层就成"天书"了。可读性 > 炫技。**一行能讲清楚就用，讲不清就老老实实写 for**。""",
      takeaway=r"""列表推导式 `[f(x) for x in 列表 if 条件]` 一行完成"加工+过滤"；lambda 是临时小函数，常当排序钥匙。**偷懒可以，但别写成只有自己看得懂的天书**。""",
      words=[
          {'en': 'COMPREHENSION', 'zh': '推导式：一行生成/加工列表的语法糖，如 [x*2 for x in a]'},
          {'en': 'LAMBDA', 'zh': 'lambda：匿名小函数，常用于排序/筛选的 key'},
          {'en': 'READABLE', 'zh': '可读性：代码让人看懂比炫技更重要'},
      ],
      code=r"""prices = ['39', '25', '暂无', '12']
nums = [int(p) for p in prices if p.isdigit()]
print('干净的数字:', nums)
books = [{'t': 'B', 'p': 25}, {'t': 'A', 'p': 30}, {'t': 'C', 'p': 20}]
books.sort(key=lambda x: x['p'])
print('按价排序:', [b['t'] for b in books])""",
      exercises=[
          {'type': 'choice', 'question': '下面哪个是列表推导式？', 'options': ['[x*2 for x in a]', 'for x in a: print(x)', 'def f(x): return x*2'], 'answer': 0, 'explain': '中括号包起来的"表达式 for 变量 in 可迭代"就是列表推导式。'},
          {'type': 'fill', 'question': 'lambda 用来定义______（填"匿名"或"有名"）小函数。', 'answer': '匿名', 'explain': 'lambda 没有名字，用完即弃，适合当 key 等场景。'},
          {'type': 'coding', 'question': '用列表推导式把下面字符串列表里的数字字符串转成整数（跳过非数字），目标输出含 [1, 2, 3]。', 'starter': "items = ['1', 'x', '2', '?', '3']\n# 一行推导式得到整数列表", 'expect': '[1, 2, 3]', 'hint': '用 [int(x) for x in items if x.isdigit()]。'},
          {'type': 'tap', 'question': '关于推导式和 lambda，说法对的有？（多选）', 'options': ['推导式能一步完成加工+过滤', 'lambda 常当 sort 的 key', '推导式套三层仍易读', '可读性比炫技重要'], 'multi': True, 'answer': [0, 1, 3], 'explain': '推导式擅长加工过滤、lambda 常作 key；但三层嵌套就难读了，可读性优先。'},
          {'type': 'open', 'question': '举一个你爬虫里"用推导式会很爽"的真实场景（比如从一堆链接里筛出图片链接）。', 'answer': '例如从所有 <a> 的 href 里筛出以 .jpg 结尾的：[h for h in hrefs if h.endswith(".jpg")]。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['COMPREHENSION', 'LAMBDA', 'READABLE']},
      ],
      tasks=[
          '在代码框里把 prices 再加几个非数字，确认推导式只留下整数。',
          '写一行推导式：从类似 ["a1","b2","c3"] 这样的字符串列表里，提取每个字符串的最后一个数字字符组成列表。',
          '用 lambda 给一组 {name, score} 按分数从高到低排序并打印名字。',
      ]),
])

# ============================ 第2章 网页与 HTTP ============================
ch2 = CH('网页与 HTTP', [
    L('r2l1', 'HTML 标签树：网页的骨架', '🌲',
      r"""## 网页本质是一棵"标签树"

你看到的漂亮页面，底层是 HTML——一堆**带尖括号的标签**，一层套一层，像一棵树。

```html
<html>
  <body>
    <h1>标题</h1>
    <p>一段文字</p>
    <a href="https://x.com">一个链接</a>
    <ul>
      <li>项目一</li>
      <li>项目二</li>
    </ul>
  </body>
</html>
```

### 三个关键概念
- **标签**：`<h1>` 标题、`<p>` 段落、`<a>` 链接、`<li>` 列表项
- **属性**：`href="..."` 是 `<a>` 的链接地址——**我们要抓的往往是它**
- **层级**：`<li>` 在 `<ul>` 里面，这种父子关系是解析时的导航图

### 爬虫怎么用
解析库（如 BeautifulSoup）把树读进来，让你说"把所有的 `<a>` 给我""把 `<h1>` 的文字给我"。**看懂树的形状，提取数据就是按图索骥**。""",
      figures=[{'key': 'html_tree', 'caption': '🌳 HTML 是一棵标签树：html→body 里套 h1/p/a/ul，ul 里再套多个 li'}],
      takeaway=r"""HTML 是一棵标签树，标签有层级、属性（如 href）藏着我们要的数据。解析库就是按这棵树来导航提取——**看懂树形，数据手到擒来**。""",
      words=[
          {'en': 'HTML', 'zh': 'HTML：网页的标记语言，用标签描述结构与内容'},
          {'en': 'TAG', 'zh': '标签：如 <p> 段落、<a> 链接，成对或自闭合'},
          {'en': 'ATTRIBUTE', 'zh': '属性：标签上的附加信息，如 <a href="..."> 里的链接地址'},
      ],
      exercises=[
          {'type': 'choice', 'question': '下面哪个标签通常表示"链接"？', 'options': ['<p>', '<a>', '<h1>'], 'answer': 1, 'explain': '<a> 是 anchor（锚点），用来做超链接，href 属性存地址。'},
          {'type': 'choice', 'question': '链接的真实地址藏在 <a> 的哪个属性里？', 'options': ['href', 'title', 'class'], 'answer': 0, 'explain': 'href 属性存放链接目标 URL。'},
          {'type': 'fill', 'question': 'HTML 标签通常成对出现，起始标签用 <标签名>，结束标签用 </___>（填符号组合）。', 'answer': '标签名', 'explain': '结束标签是 </标签名>，如 </p>；也可答"斜杠+标签名"。'},
          {'type': 'tap', 'question': '关于 HTML 标签树，说法对的有？（多选）', 'options': ['标签有父子层级', '属性可能藏着要抓的数据', '所有标签都自闭合', '解析库按树导航提取'], 'multi': True, 'answer': [0, 1, 3], 'explain': '标签有层级、属性藏数据、解析库按树提取；并非所有标签自闭合（如 <a> 成对）。'},
          {'type': 'open', 'question': '给你一段 HTML，你要抠出所有商品名（都在 <span class="name"> 里），你会怎么描述"导航路径"？写 2 句。', 'answer': '先按 class="name" 定位所有 span，再取每个的 .text；或先找到商品容器再在内部找 .name。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['HTML', 'TAG', 'ATTRIBUTE']},
      ],
      tasks=[
          '右键随便一个网页"查看网页源代码"，找 5 个 <a href=...> 看看链接长啥样。',
          '在开发者工具 Elements 面板里，点开一个列表，数数它嵌套了几层标签。',
          '画一张你目标页面的"关键数据在哪些标签/属性里"的地图（例如：标题在 <h1>，价格在一个 class="price" 的 div）。',
      ]),

    L('r2l2', '开发者工具：看网页背后的真实请求', '🛠️',
      r"""## 真正的数据，常常藏在"网络"里

按 **F12**（或右键"检查"）打开开发者工具，你会看到网页的真实内幕。

### 三个最常用面板
- **Elements（元素）**：渲染后的 HTML 树，右键元素可"复制选择器"，后面写提取代码超方便
- **Network（网络）**：记录浏览器发出的**每一个请求**。很多网页数据不是写在 HTML 里，而是由 JS 通过**接口（API）**单独拉取的——在这里能抓到原始 JSON
- **Console（控制台）**：看报错、试小段代码

### 爬虫视角怎么用
1. 打开 Network，刷新页面
2. 在请求列表里找返回数据是 JSON 或大片 HTML 的那条
3. 看它的 **URL、请求方法、请求头、响应内容**
4. 把这些信息搬进你的 `requests` 代码

> 经验铁律：**先用 Network 看"数据到底从哪个请求来"**，再决定抓 HTML 还是直接打接口，能省一大半力气。""",
      takeaway=r"""F12 开发者工具是你的"透视镜"。Network 面板能看见网页背后的真实请求与接口。**写代码前先按 F12 看 Network，搞清楚数据从哪来**——想清楚这步，后面少走三天弯路。""",
      words=[
          {'en': 'DEVTOOLS', 'zh': '开发者工具：浏览器内置审查面板，按 F12 打开'},
          {'en': 'NETWORK', 'zh': 'Network 面板：记录所有网络请求，找接口和数据的入口'},
          {'en': 'CONSOLE', 'zh': 'Console 控制台：看报错、试小段代码的地方'},
      ],
      exercises=[
          {'type': 'choice', 'question': '想看网页到底向哪些地址发了请求、返回了什么，该用哪个面板？', 'options': ['Network', 'Elements', 'Console'], 'answer': 0, 'explain': 'Network 面板记录全部请求与响应，是找数据源的第一现场。'},
          {'type': 'choice', 'question': '很多网页的数据其实是由 JS 通过什么单独拉取的？', 'options': ['接口(API)', '打印机', '显卡'], 'answer': 0, 'explain': '现代网页常由 JS 调接口拿 JSON，不在初始 HTML 里。'},
          {'type': 'fill', 'question': '在开发者工具里，记录浏览器所有网络请求、用来找数据的面板叫 ______。', 'answer': 'network', 'explain': 'Network 面板是定位数据源的关键。'},
          {'type': 'tap', 'question': '用开发者工具定位数据源的正确姿势包括？（多选）', 'options': ['F12 打开工具', '切到 Network 刷新', '在请求列表找返回 JSON/HTML 的那条', '盲猜一个 URL 就开爬'], 'multi': True, 'answer': [0, 1, 2], 'explain': '先开工具、切 Network、刷新、再定位数据请求；盲猜 URL 是新手最容易踩的坑。'},
          {'type': 'open', 'question': '你目标网站首页看起来有数据，但"查看源代码"里却找不到——你怎么判断它是 HTML 里写的还是接口拉的？写 2-3 句。', 'answer': '先看页面源代码搜关键词，没有就去 F12 的 Network 面板找 XHR/Fetch 请求，看哪个响应里带数据；基本能判定是接口动态加载。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['DEVTOOLS', 'NETWORK', 'CONSOLE']},
      ],
      tasks=[
          '打开 B 站/豆瓣任意一个列表页，F12 → Network，刷新，找一条返回 JSON 的 XHR 请求，把它的 URL 和"响应"前几行截图或抄下来。',
          '在 Elements 面板里右键一个链接，点"复制选择器"，看看生成的选择器长啥样。',
          '用 Console 敲 document.querySelectorAll("a").length，看当前页有多少个链接——感受一下“DOM 里能数出来”。',
      ]),

    L('r2l3', 'URL 与 请求/响应：按门铃取货', '🔗',
      r"""## URL 是"门牌号"，HTTP 是"取货协议"

### 拆开一个 URL
```
https://www.example.com/search?q=python&page=2
│││  ││││││││││││││  │││││││││││││  └ 查询参数 q=python&page=2
协议  主机/域名          路径 /search
```
- **协议** `https`：加密的 HTTP
- **主机** `www.example.com`：哪台服务器
- **路径** `/search`：要哪个资源
- **查询参数** `?q=python&page=2`：给服务器的附加条件，多个用 `&` 连

### 请求与响应
- **请求(Request)**：你发出去，含 方法(GET/POST)、URL、请求头(Headers)、可选的参数/正文
- **响应(Response)**：服务器回来，含 状态码、响应头、**正文**（HTML 或 JSON）

爬虫就是**构造一个"请求"，再读取"响应"的正文**。下一章用 `requests` 真发请求。""",
      figures=[{'key': 'request_response', 'caption': '📨 请求=方法+URL+头+参数；响应=状态码+头+正文(HTML/JSON)。爬虫构造请求、读取正文'}],
      takeaway=r"""URL 拆成 协议/主机/路径/参数；HTTP 是"请求→响应"模型，**响应正文就是你要抓的 HTML 或 JSON**。把门牌拆清楚、按门铃姿势摆对，数据才乖乖递出来。""",
      words=[
          {'en': 'URL', 'zh': 'URL：统一资源定位符，即网址'},
          {'en': 'QUERY', 'zh': '查询参数：URL 里 ? 后面的 k=v，多个用 & 连接'},
          {'en': 'RESPONSE', 'zh': '响应：服务器收到请求后返回的内容'},
      ],
      exercises=[
          {'type': 'choice', 'question': 'URL 中 `?q=python&page=2` 这部分叫？', 'options': ['查询参数', '主机名', '协议'], 'answer': 0, 'explain': '问号后是查询参数，多个键值对用 & 连接。'},
          {'type': 'choice', 'question': '服务器返回的内容（HTML/JSON）属于？', 'options': ['请求', '响应正文', 'URL'], 'answer': 1, 'explain': '响应里的正文承载真正的数据。'},
          {'type': 'fill', 'question': 'URL 里多个查询参数之间用 ___ 符号连接。', 'answer': '&', 'explain': '多个参数用 & 连接，如 ?a=1&b=2。'},
          {'type': 'tap', 'question': '一个完整 URL 通常包含哪些部分？（多选）', 'options': ['协议(https)', '主机/域名', '路径', '查询参数'], 'multi': True, 'answer': [0, 1, 2, 3], 'explain': '协议、主机、路径、查询参数四件套构成常见 URL。'},
          {'type': 'open', 'question': '你目标站点翻到第 3 页时，地址栏 URL 通常怎么变？你打算怎么用这个规律做"自动翻页"？写 2 句。', 'answer': '翻页往往体现在 page=3 或 offset=20 这类参数上；自动翻页就是循环改变这个参数、每次重新请求。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['URL', 'QUERY', 'RESPONSE']},
      ],
      tasks=[
          '复制一个带 ? 参数的真实网址，手动拆成"协议/主机/路径/参数"四段写下来。',
          '在你目标站点翻两页，对比两个 URL，找出"页码"藏在哪个参数里。',
          '用纸笔写出"请求"和"响应"各自包含哪几样东西（方法/URL/头/正文等）。',
      ]),

    L('r2l4', '状态码与编码：读懂服务器的表情', '🔢',
      r"""## 状态码：服务器用三位数"回话"

每次响应都带一个**状态码**，告诉你"要的东西怎么样了"：
- `200` 成功，一切正常
- `301/302` 跳转，地址变了（爬虫要跟着重定向）
- `403` 禁止访问（常被反爬拦下，服务器：你谁啊）
- `404` 找不到页面
- `500` 服务器自己崩了
- `503` 服务不可用（可能限流，稍后重试）

### 编码：别让中文变乱码
网页正文可能是 `utf-8`、`gbk` 等。`requests` 一般会猜对，但遇到乱码，手动指定最稳：
```python
r = requests.get(url)
r.encoding = 'utf-8'      # 或 'gbk'
text = r.text
```

### 小结
看到 `200` 再处理；遇到 `4xx/5xx` 先排查（是不是被拦、地址对不对）。**编码统一用 utf-8，是爬虫不出乱码的铁律**。""",
      figures=[{'key': 'status_wheel', 'caption': '🔢 状态码速记：2xx成功 / 3xx跳转 / 4xx你的问题(403被拦) / 5xx服务器抽风'}],
      takeaway=r"""状态码 200 成功、4xx 客户端问题（403=被反爬拦）、5xx 服务器问题；**中文乱码就手动设 r.encoding='utf-8'**，编码统一最稳。""",
      words=[
          {'en': 'STATUS', 'zh': '状态码：响应里三位数，表明请求结果（200成功等）'},
          {'en': 'ENCODING', 'zh': '编码：字符如何变成字节，中文常用 utf-8/gbk'},
          {'en': 'REDIRECT', 'zh': '重定向：301/302，地址跳转，爬虫要跟随'},
      ],
      exercises=[
          {'type': 'choice', 'question': '状态码 200 表示？', 'options': ['请求成功', '页面不存在', '服务器错误'], 'answer': 0, 'explain': '2xx 是成功，200 最常见。'},
          {'type': 'choice', 'question': '403 通常意味着？', 'options': ['服务器说"你谁啊，禁止访问"', '页面搬家了', '请求成功'], 'answer': 0, 'explain': '403 Forbidden，常被反爬用来拦爬虫。'},
          {'type': 'fill', 'question': '遇到中文乱码，最稳的做法是设 r.encoding = ___ 。', 'answer': "'utf-8'", 'explain': "手动指定编码可消除大部分乱码；也可答 utf-8 不带引号。"},
          {'type': 'tap', 'question': '关于状态码，说法对的有？（多选）', 'options': ['2xx 成功', '301/302 是跳转', '404 找不到页面', '500 是服务器自己崩'], 'multi': True, 'answer': [0, 1, 2, 3], 'explain': '四类都正确：2xx成功、3xx跳转、404丢失、5xx服务器错。'},
          {'type': 'open', 'question': '你爬虫连发请求，突然开始大量收到 503，你第一反应该排查什么、怎么应对？写 2-3 句。', 'answer': '503 多半是限流/过载；先降速、加重试与退避，确认是否被封，必要时换 IP 或暂停，别硬刚把人家打挂。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['STATUS', 'ENCODING', 'REDIRECT']},
      ],
      tasks=[
          '用 `curl -I 一个网址` 或浏览器看响应头，确认它返回的状态码是多少。',
          '找一个 gbk 编码的老网站（如某些政府/老论坛），把编码手动设成 gbk 再试抓一次，看是否还乱码。',
          '把"2xx/3xx/4xx/5xx 各自代表啥"写成一张小抄贴桌上。',
      ]),
])

# ============================ 第3章 requests 抓取 ============================
ch3 = CH('requests 抓取', [
    L('r3l1', '第一个 GET 请求：三行搬回网页', '📡',
      r"""## 三行代码，把网页搬回家

`requests` 是爬虫发请求的首选库。最简单的抓取长这样：

```python
import requests

url = 'https://www.example.com'
r = requests.get(url)
print(r.status_code)     # 200
print(r.text[:200])      # 网页前 200 个字符
```

### 逐行拆解
- `requests.get(url)`：发出一个 **GET** 请求，返回**响应对象** `r`
- `r.status_code`：状态码，先看它是不是 200
- `r.text`：响应正文（HTML 字符串），这就是后面要解析的原材料

### 拿到手先别急着解析
养成习惯：**先打印 status_code 和 text 前几百字**，确认拿到了正常 HTML，再往下写。很多人一上来就解析，结果解析了个寂寞（拿到的是空页或反爬页）。

> 浏览器里的迷你 Python 跑不了 requests（要联网+装库），所以这一节**代码在你自己电脑上跑**。下面的"任务"就是让你真跑一遍。""",
      takeaway=r"""requests.get(url) 一行发出 GET 请求，r.status_code 看成败、r.text 拿 HTML 正文。**拿到手先确认 200 和正文正常，再解析**——别还没看清数据就上手抠。""",
      words=[
          {'en': 'GET', 'zh': 'GET：最常用的请求方法，用来"取"资源'},
          {'en': 'RESPONSE', 'zh': '响应对象：requests 拿到的结果，含状态码和正文'},
          {'en': 'TEXT', 'zh': 'r.text：响应正文，HTML 或 JSON 的字符串'},
      ],
      exercises=[
          {'type': 'choice', 'question': '`requests.get(url)` 返回的是什么？', 'options': ['响应对象 r', '一段 HTML 文本', '状态码数字'], 'answer': 0, 'explain': 'get 返回响应对象，状态码和正文都要从它身上取。'},
          {'type': 'choice', 'question': '拿到响应后，第一步建议先做什么？', 'options': ['打印 status_code 和 text 前若干字确认正常', '立刻写解析逻辑', '直接保存文件'], 'answer': 0, 'explain': '先确认 200 且正文正常，避免解析空页/反爬页。'},
          {'type': 'fill', 'question': '想拿响应的 HTML 正文，应访问 r.______ 属性。', 'answer': 'text', 'explain': 'r.text 是响应正文字符串。'},
          {'type': 'order', 'question': '用 requests 抓一个网页的正确顺序：', 'steps': ['import requests', 'r = requests.get(url)', '检查 r.status_code', '读取 r.text 解析'], 'explain': '导入→发请求→看状态码→读正文，顺序别乱。'},
          {'type': 'tap', 'question': '关于 requests.get，说法对的有？（多选）', 'options': ['它发出 GET 请求', '返回对象里有 status_code', '返回对象里有 text', '它只能发 POST'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'get 发 GET、返回对象含 status_code 和 text；发 POST 用 requests.post。'},
          {'type': 'open', 'question': '你拿到的 r.text 开头是 "<!DOCTYPE html>" 说明什么？如果是空字符串又可能说明什么？写 2-3 句。', 'answer': '以 DOCTYPE 开头说明正常拿到了 HTML 页面；空字符串可能是被反爬拦截、需要登录、或请求方式/参数不对。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['GET', 'RESPONSE', 'TEXT']},
      ],
      tasks=[
          '在自己电脑跑一遍上面的三段代码（把 url 换成你常逛的站点），看 status_code 和 text 前 200 字。',
          '故意访问一个不存在的页面（如 example.com/nope），观察 status_code 变成多少。',
          '把 r.text 前 500 字保存到 review.html，用浏览器打开，确认那就是网页源码。',
      ]),

    L('r3l2', 'Headers 与 UA 伪装：别一上来就被认成机器人', '🎭',
      r"""## 服务器怎么认出你是爬虫？

默认 `requests` 的请求头里，`User-Agent` 写的是 `python-requests/xx`——等于举着牌子喊"我是爬虫"。很多站点一看就拦。

### 伪装成浏览器
```python
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36',
    'Accept-Language': 'zh-CN,zh;q=0.9',
}
r = requests.get(url, headers=headers)
```

### 还有几个常用头
- `Referer`：告诉服务器"我从哪个页面点过来的"，有些接口会校验
- `Cookie`：带登录态（第 7 章细讲）
- `Accept`：声明我能接受什么类型的数据

### 心态
伪装 UA 不是"欺骗"，是**礼貌地表明你是个正常浏览器**，避免被误伤。但别伪装成别家公司或伪造身份干坏事——那是另一回事。""",
      takeaway=r"""默认 requests 的 UA 写着"我是爬虫"容易被拦。发请求时带上浏览器般的 headers（尤其 User-Agent）更稳。**伪装 UA 是礼貌自报身份，不是伪造去干坏事**。""",
      words=[
          {'en': 'HEADER', 'zh': '请求头：请求里附带的自述信息，如 User-Agent'},
          {'en': 'USERAGENT', 'zh': 'User-Agent：表明"我是谁/什么客户端"的字段'},
          {'en': 'REFERER', 'zh': 'Referer：表明"从哪个页面过来"，部分接口会校验'},
      ],
      exercises=[
          {'type': 'choice', 'question': '默认 requests 的 User-Agent 通常会暴露什么？', 'options': ['python-requests 字样，等于自报爬虫', 'Chrome 浏览器', '什么都没有'], 'answer': 0, 'explain': '默认 UA 含 python-requests，容易被反爬识别拦截。'},
          {'type': 'fill', 'question': '想伪装成浏览器，要在 get 时传入 ______ 参数（填参数名）。', 'answer': 'headers', 'explain': 'requests.get(url, headers={...}) 传入请求头字典。'},
          {'type': 'tap', 'question': '以下哪些常作为请求头字段使用？（多选）', 'options': ['User-Agent', 'Referer', 'Cookie', 'status_code'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'UA/Referer/Cookie 都是常见请求头；status_code 是响应里的，不是请求头。'},
          {'type': 'choice', 'question': '关于"伪装 UA"，下面哪种说法正确？', 'options': ['是礼貌自报身份避免误伤，不是伪造去作恶', '随便伪装成别家公司都行', '完全没必要'], 'answer': 0, 'explain': '伪装成正常浏览器可避免被误拦，但伪造身份作恶是另一回事。'},
          {'type': 'open', 'question': '你爬一个站点总是 403，怀疑是 UA 被拦。你打算怎么验证并解决？写 2-3 句。', 'answer': '先用默认 UA 请求看是否 403，再带上真实浏览器 UA 重试；若仍 403 再排查 Referer/Cookie 或反爬机制。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['HEADER', 'USERAGENT', 'REFERER']},
      ],
      tasks=[
          '用默认 UA 请求一个站点，打印 r.request.headers["User-Agent"] 看看它写了啥。',
          '加上一段浏览器 UA 再请求同站点，对比两次的 status_code / 正文长度。',
          '从自己浏览器的开发者工具 Network 里复制一个真实 User-Agent 字符串，存成变量备用。',
      ]),

    L('r3l3', '参数与 POST：GET 带参 vs 表单提交', '📨',
      r"""## 两种"递纸条"的方式

### GET：把参数挂在 URL 上
翻页、搜索关键词，常常就是改 URL 的查询参数：
```python
params = {'q': 'python', 'page': 2}
r = requests.get('https://x.com/search', params=params)
# 实际请求 https://x.com/search?q=python&page=2
```
用 `params` 比手写字符串拼接安全（自动编码特殊字符）。

### POST：把参数塞进请求体
登录、提交表单，数据不走 URL，藏在**请求体**里：
```python
data = {'username': 'me', 'password': '123'}
r = requests.post('https://x.com/login', data=data)
```
> 密码千万别写进 URL（会进浏览器历史/服务器日志），该 POST 就 POST。

### 怎么判断用哪个
打开开发者工具 Network，看那条请求的**方法**是 GET 还是 POST，照抄它的参数和格式，基本就通了。""",
      takeaway=r"""GET 用 params 把参数挂 URL（自动编码，适合搜索/翻页）；POST 用 data 把参数塞请求体（适合登录/提交，密码别进 URL）。**先看 Network 里目标是 GET 还是 POST，照抄格式**。""",
      words=[
          {'en': 'PARAMS', 'zh': 'params：GET 请求的查询参数，requests 会自动拼到 URL'},
          {'en': 'POST', 'zh': 'POST：把数据放请求体提交的请求方法'},
          {'en': 'FORM', 'zh': '表单：网页里收集用户输入再提交的结构，常走 POST'},
      ],
      exercises=[
          {'type': 'choice', 'question': '用 requests 发带查询参数的 GET，正确写法是？', 'options': ['requests.get(url, params={"q":1})', 'requests.get(url+"?q=1") 手写', 'requests.post(url, params={"q":1})'], 'answer': 0, 'explain': 'get 配合 params 字典，requests 自动编码拼接，最稳。'},
          {'type': 'choice', 'question': '登录账号密码一般该用哪种方式提交？', 'options': ['POST 放请求体', 'GET 放 URL', '随便'], 'answer': 0, 'explain': '密码走 POST 请求体，避免进 URL 历史/日志泄露。'},
          {'type': 'fill', 'question': 'requests.post(url, ______={"username":"me"}) 里该填的参数名是 data（提交表单体）。', 'answer': 'data', 'explain': 'POST 表单数据用 data= 传入。'},
          {'type': 'tap', 'question': '关于 GET 与 POST，说法对的有？（多选）', 'options': ['GET 参数在 URL 上', 'POST 参数在请求体', '密码应走 POST', 'GET 比 POST 更安全'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'GET 参数在 URL、POST 在体；密码走 POST；"GET 更安全"是误解，安全靠 HTTPS 不靠方法。'},
          {'type': 'open', 'question': '你在 Network 里看到目标请求是 POST，但参数格式看不懂（有 token、有嵌套）。你打算怎么稳妥地复现它？写 2-3 句。', 'answer': '复制该请求为 curl（Network 里右键→Copy as cURL），照着它的 URL、方法、headers、data 逐个搬到 requests；token 类参数先原样带入再研究来源。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['PARAMS', 'POST', 'FORM']},
      ],
      tasks=[
          '用 params 请求一个搜索接口（如某站内搜），打印最终 r.url 看参数是否拼对。',
          '找一个登录页，用 POST 提交一次（可用测试账号），看响应里有没有登录成功的标志。',
          '在 Network 里右键一条请求"Copy as cURL"，对比它和你的 requests 代码差在哪。',
      ]),

    L('r3l4', '超时、重试与会话：网络不是你家，会抽风', '🔌',
      r"""## 真实网络：没有 100% 可靠

你电脑→服务器的链路长着呢，随时可能卡、断、慢。爬虫必须**为失败做准备**。

### 超时：别无限等
```python
r = requests.get(url, timeout=10)   # 10 秒还没响应就抛异常
```
不设超时，一个卡死的请求能让你脚本挂一整晚。

### 重试：失败就再来一次
```python
from time import sleep
for i in range(3):
    try:
        r = requests.get(url, timeout=10)
        break
    except Exception as e:
        print('第', i+1, '次失败，睡 2 秒重试')
        sleep(2)
```

### 会话：保持"同一个我"
连续请求同一站点（比如先登录再抓），用 `requests.Session()` 自动**复用 Cookie 和连接**，比每次 get 都新建干净利落：
```python
s = requests.Session()
s.get('https://x.com/login', data=...)   # 登录，Cookie 存进 s
s.get('https://x.com/vip-data')          # 带着登录态访问
```""",
      figures=[{'key': 'session_cookie', 'caption': '🍪 Session 复用同一 Cookie 与连接：登录一次，后续请求自动带登录态，比每次新建干净'}],
      takeaway=r"""真实网络会抽风：**超时(timeout)防卡死、重试(for+except+sleep)防偶发失败、Session 复用 Cookie 保持登录态**。把"可能失败"写进代码，爬虫才扛造。""",
      words=[
          {'en': 'TIMEOUT', 'zh': '超时：请求等待的最长时间，到了就放弃，防卡死'},
          {'en': 'RETRY', 'zh': '重试：失败后隔会儿再试一次，应对网络抖动'},
          {'en': 'SESSION', 'zh': '会话：用 Session() 复用 Cookie 与连接，保持登录态'},
      ],
      exercises=[
          {'type': 'choice', 'question': '给请求加 timeout=10 的作用是？', 'options': ['10 秒没响应就放弃，防卡死', '限制下载速度', '限制重试次数'], 'answer': 0, 'explain': '超时避免一个请求无限挂起拖垮整个脚本。'},
          {'type': 'choice', 'question': '连续请求（先登录再抓）用什么保持登录态最方便？', 'options': ['requests.Session()', '每次都重新 get', '把密码写进 URL'], 'answer': 0, 'explain': 'Session 自动复用 Cookie 与连接，登录一次后续自动带态。'},
          {'type': 'fill', 'question': '应对偶发网络失败，常用 for 循环 + ___ 捕获异常 + sleep 实现重试。', 'answer': 'try', 'explain': 'try/except 包住请求，失败 sleep 后重试。'},
          {'type': 'order', 'question': '一个稳健请求的典型顺序：', 'steps': ['设置 timeout', 'try 发请求', '失败 except 后 sleep', '重试若干次'], 'explain': '先设超时，再 try 发请求，失败 sleep 重试，循环若干次。'},
          {'type': 'tap', 'question': '关于健壮性，说法对的有？（多选）', 'options': ['该设超时', '可加重试', 'Session 能保持登录态', '网络永远可靠不用管'], 'multi': True, 'answer': [0, 1, 2], 'explain': '超时、重试、Session 都提升健壮性；"网络永远可靠"是幻觉。'},
          {'type': 'open', 'question': '你爬虫半夜跑，白天发现卡在某一页 8 小时没动。最可能是漏了什么设置？怎么改？写 2-3 句。', 'answer': '漏了 timeout，导致请求无限挂起；加上 timeout=10 并用 try/except 兜底，失败则跳过或重试该页。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['TIMEOUT', 'RETRY', 'SESSION']},
      ],
      tasks=[
          '给之前写的请求加上 timeout=10，故意请求一个很慢/不存在的地址，看是否如期抛超时。',
          '写一个简单的重试函数：失败最多 3 次，每次间隔 2 秒。',
          '用 Session 连续请求同一站点两次，打印第二次请求自动带上的 Cookie（s.cookies）。',
      ]),

    L('r3l5', '状态码与错误处理：见到 4xx/5xx 别慌', '🚦',
      r"""## 把"失败"当正常流程的一部分

抓一百个页面，不可能个个 200。成熟的爬虫会**按状态码分流处理**。

### 常见套路
```python
r = requests.get(url, timeout=10)
if r.status_code == 200:
    parse(r.text)                 # 正常解析
elif r.status_code == 404:
    print('页面丢了，跳过')         # 不存在，记日志跳过
elif r.status_code == 403:
    print('被拦了，可能要换 UA/代理') # 反爬，换策略
elif r.status_code >= 500:
    print('服务器抽风，稍后重试')     # 服务端问题，退避重试
else:
    print('其他状态', r.status_code)
```

### 几条经验
- **403**：先换 UA、加 Referer；还不行再看是否需要登录/代理
- **429 / 503**：被限流了，降速 + 退避重试
- **3xx**：requests 默认自动跟随重定向，一般不用管；要禁就 `allow_redirects=False`

> 把"坏响应"也写成代码分支，你的爬虫才不会因为一个 404 全体罢工。""",
      takeaway=r"""按状态码分流：200 解析、404 跳过、403 换 UA/策略、5xx 退避重试。**把坏响应也写成分支**，爬虫才不会因为单个 404 全体罢工。""",
      words=[
          {'en': 'STATUS', 'zh': '状态码：响应里三位数，表明请求结果'},
          {'en': 'REDIRECT', 'zh': '重定向：3xx，requests 默认自动跟随'},
          {'en': 'THROTTLE', 'zh': '限流：429/503 表示被限速，需降速退避'},
      ],
      exercises=[
          {'type': 'choice', 'question': '收到 403，第一反应通常该？', 'options': ['换 UA/加 Referer 或检查是否需登录', '立刻狂刷重试', '直接放弃整个爬虫'], 'answer': 0, 'explain': '403 常见于被反爬拦，先换 UA/头或确认登录态。'},
          {'type': 'choice', 'question': '429 或 503 通常意味着？', 'options': ['被限流了，该降速退避', '页面不存在', '请求方法错了'], 'answer': 0, 'explain': '429/503 多为限流或过载，需降速、加重试退避。'},
          {'type': 'fill', 'question': 'requests 默认会自动跟随 ___ 状态码（填 2xx/3xx/4xx/5xx 之一）。', 'answer': '3xx', 'explain': '3xx 重定向默认自动跟随；要禁设 allow_redirects=False。'},
          {'type': 'tap', 'question': '关于状态码处理，做法对的有？（多选）', 'options': ['200 才解析', '404 记日志跳过', '5xx 退避重试', '所有非 200 都立刻崩溃'], 'multi': True, 'answer': [0, 1, 2], 'explain': '按码分流是成熟做法；"非200就崩"是新手写法。'},
          {'type': 'open', 'question': '你爬虫抓 1000 页，其中 30 页返回 404。你希望脚本怎么表现才合理？写 2-3 句。', 'answer': '遇到 404 记日志并跳过该页，继续抓其余页面，最后汇总失败列表；不应因个别 404 中断整体。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['STATUS', 'REDIRECT', 'THROTTLE']},
      ],
      tasks=[
          '写一段"按状态码分流"的代码骨架（if/elif 覆盖 200/403/404/5xx），用假数据走一遍。',
          '故意请求一个 404 地址，确认你的分支能识别并跳过而非崩溃。',
          '统计你之前抓的一批 URL 的状态码分布，看看健康比例多少。',
      ]),
])

# ============================ 第4章 BeautifulSoup 解析 ============================
ch4 = CH('BeautifulSoup 解析', [
    L('r4l1', '解析 HTML：把字符串变成可导航的树', '🌳',
      r"""## 拿到 HTML 字符串后，下一步是"读懂它"

`BeautifulSoup`（简称 bs4）把一段 HTML 字符串解析成**可导航的对象树**，你就能用 Python 语法去捞数据，而不是用正则硬抠（正则抠 HTML 是出了名的痛苦）。

### 起手式
```python
from bs4 import BeautifulSoup
import requests

html = requests.get(url, headers=headers).text
soup = BeautifulSoup(html, 'html.parser')   # 或 'lxml'（更快，需 pip install lxml）
```

### 两种解析器
- `html.parser`：Python 自带，零依赖，够用
- `lxml`：快、容错好，但要额外装

### 它解决了什么
正则匹配 HTML 容易翻车（标签嵌套、属性顺序乱）。bs4 按**树结构**理解文档：你能说"找所有 `<a>`""找 class 为 x 的 div""取某个标签的文字"，稳得多。

> 下一节正式开捞。这一节先把 `soup = BeautifulSoup(html, 'html.parser')` 这行刻进手。""",
      takeaway=r"""BeautifulSoup 把 HTML 字符串变成可导航的对象树，比正则硬抠稳得多。起手就是 `soup = BeautifulSoup(html, 'html.parser')`（要快就装 lxml 用 'lxml'）。""",
      words=[
          {'en': 'BEAUTIFULSOUP', 'zh': 'BeautifulSoup(bs4)：把 HTML 解析成对象树、方便提取的库'},
          {'en': 'SOUP', 'zh': 'soup：BeautifulSoup 解析后的文档对象，用来导航查找'},
          {'en': 'PARSER', 'zh': '解析器：把文本变成结构的东西，如 html.parser / lxml'},
      ],
      exercises=[
          {'type': 'choice', 'question': 'BeautifulSoup 相比"用正则抠 HTML"的最大优势是？', 'options': ['按树结构理解文档，更稳', '更快', '不需要装库'], 'answer': 0, 'explain': 'HTML 嵌套复杂，正则易翻车；bs4 按树导航稳得多。'},
          {'type': 'choice', 'question': '下面哪行能正确创建 soup 对象？', 'options': ["BeautifulSoup(html, 'html.parser')", 'Soup(html)', 'bs4(html)'], 'answer': 0, 'explain': '标准起手：BeautifulSoup(html, 解析器名)。'},
          {'type': 'fill', 'question': '想用更快的 lxml 解析器，创建时应写 BeautifulSoup(html, ___ )。', 'answer': "'lxml'", 'explain': "传 'lxml' 需先 pip install lxml。"},
          {'type': 'tap', 'question': '关于 bs4 解析器，说法对的有？（多选）', 'options': ['html.parser 自带零依赖', 'lxml 更快需另装', '解析器把文本变结构', '必须联网才能解析'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'html.parser 自带、lxml 更快需装、解析器产出结构；解析本地字符串不需联网。'},
          {'type': 'open', 'question': '为什么"用正则抠 HTML"容易翻车？举一个会翻车的例子（比如标签嵌套或属性顺序）。写 2 句。', 'answer': 'HTML 标签可嵌套、属性顺序不固定、还可能缺引号，正则难以覆盖所有变体；如 <div class="a" id="b"> 与 <div id="b" class="a"> 正则要写两种，bs4 一个选择器搞定。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['BEAUTIFULSOUP', 'SOUP', 'PARSER']},
      ],
      tasks=[
          '用 requests 抓一个真实页面存成 html 字符串，再用 BeautifulSoup 解析出 soup。',
          '比较 html.parser 与 lxml（装一下）解析同一个页面，看速度差异。',
          '打印 soup.title 和 soup.title.string，看能不能拿到 <title> 内容。',
      ]),

    L('r4l2', 'find / find_all：精准捞数据', '🥄',
      r"""## 最常用的两把勺子

- `find(name, attrs)`：返回**第一个**匹配的标签
- `find_all(name, attrs)`：返回**所有**匹配的标签（列表）

```python
# 拿第一个 <h1> 的文字
print(soup.find('h1').text)

# 拿所有链接
links = soup.find_all('a')
for a in links:
    print(a.get('href'))      # 取 href 属性

# 按 class 找（class 是关键字，加下划线）
prices = soup.find_all('span', class_='price')
for p in prices:
    print(p.text)
```

### 几个坑
- **取文字用 `.text`**（或 `.get_text()`），别直接 print 标签（会带 HTML）
- **取属性用 `.get('href')`**，比 `a['href']` 安全（没有该属性时 `.get` 返回 None 不报错）
- `class_` 要加下划线，因为 `class` 是 Python 关键字

> 小光口诀：**find 拿一个，find_all 拿一堆；.text 取字，.get 取属性。**""",
      takeaway=r"""find 拿第一个、find_all 拿所有；**.text 取文字、.get('href') 取属性**；按 class 找要写 class_（因为 class 是关键字）。口诀：find 一个 find_all 一堆，text 取字 get 取属性。""",
      words=[
          {'en': 'FIND', 'zh': 'find：返回第一个匹配的标签'},
          {'en': 'FINDALL', 'zh': 'find_all：返回所有匹配的标签列表'},
          {'en': 'GETTEXT', 'zh': '.text / .get_text()：取标签里的文字；.get() 取属性'},
      ],
      exercises=[
          {'type': 'choice', 'question': '想拿到页面里**所有**的 <a> 链接，用？', 'options': ['find_all("a")', 'find("a")', 'get("a")'], 'answer': 0, 'explain': 'find_all 返回全部匹配，find 只返回第一个。'},
          {'type': 'choice', 'question': '取一个标签里的文字，正确写法是？', 'options': ['.text', '.string 永远更好', '.href'], 'answer': 0, 'explain': '.text 取标签内全部文字；.href 是属性不是文字。'},
          {'type': 'fill', 'question': '按 class 查找时要写 class_（下划线），因为 class 是 Python 的______。', 'answer': '关键字', 'explain': 'class 是保留字，bs4 用 class_ 避开。'},
          {'type': 'tap', 'question': '关于取属性，说法对的有？（多选）', 'options': ['用 .get("href") 更安全', 'a["href"] 没有该属性会报错', '.text 取文字', 'find 返回列表'], 'multi': True, 'answer': [0, 1, 2], 'explain': '.get 缺属性返回 None 更稳；a["href"] 缺失会 KeyError；.text 取字；find 只返回一个。'},
          {'type': 'open', 'question': '你要抠 50 个商品的价格，每个都在 <span class="price">9.9</span> 里。写 3 行核心代码思路（不用写完整）。', 'answer': 'prices = soup.find_all("span", class_="price"); for p in prices: print(p.text)。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['FIND', 'FINDALL', 'GETTEXT']},
      ],
      tasks=[
          '用 find_all("a") 抠出你目标页面的所有链接，打印前 20 个 href。',
          '按 class_="price"（或你页面真实类名）抠出所有价格并打印。',
          '对比 .text 与直接打印标签的区别：打印一个 <p> 标签本身 vs 它的 .text。',
      ]),

    L('r4l3', 'CSS 选择器：用 .class #id tag 定位', '🎯',
      r"""## 当你想"像前端那样选元素"

`select()` 让你用 **CSS 选择器** 定位，和开发者工具里"复制选择器"拿到的格式一致，上手极快。

```python
# 所有 class 为 price 的元素
soup.select('.price')

# id 为 main 的 div 下的所有 p
soup.select('div#main p')

# 多层：nav 里的所有链接
soup.select('nav a')

# 取第一个
soup.select_one('.title')
```

### 选择器速查
| 写法 | 含义 |
|---|---|
| `tag` | 选某种标签，如 `p` |
| `.cls` | 选 class 含 cls |
| `#id` | 选 id 为 id |
| `a b` | b 是 a 的后代 |
| `a > b` | b 是 a 的直接子元素 |

### 什么时候用 select
开发者工具"复制 selector"得到的就是 CSS 选择器，直接粘进 `select()` 最省事。**find/find_all 和 select 二选一都行，看哪个顺手**。""",
      takeaway=r"""select() 用 CSS 选择器定位，和开发者工具"复制选择器"同款，最省事。`.cls` 选类、`#id` 选 id、`a b` 选后代。find 和 select 二选一，顺手就行。""",
      words=[
          {'en': 'SELECT', 'zh': 'select：用 CSS 选择器查找元素，返回列表'},
          {'en': 'SELECTOR', 'zh': '选择器：如 .class / #id / tag 的定位表达式'},
          {'en': 'SELECTONE', 'zh': 'select_one：用选择器取第一个匹配'},
      ],
      exercises=[
          {'type': 'choice', 'question': 'CSS 选择器里 `.price` 表示？', 'options': ['class 为 price 的元素', 'id 为 price', '标签 price'], 'answer': 0, 'explain': '点号开头是 class 选择器。'},
          {'type': 'choice', 'question': '`soup.select_one("#title")` 返回？', 'options': ['id 为 title 的第一个元素', '所有 id 为 title', '报错'], 'answer': 0, 'explain': '# 是 id 选择器，select_one 取第一个。'},
          {'type': 'fill', 'question': 'CSS 选择器里 `div p` 表示 p 是 div 的______（填"后代"或"兄弟"）。', 'answer': '后代', 'explain': '空格表示后代关系（不一定是直接子元素）。'},
          {'type': 'tap', 'question': '关于 CSS 选择器，说法对的有？（多选）', 'options': ['.cls 选 class', '#id 选 id', 'a b 选后代', '只能选标签不能选 class'], 'multi': True, 'answer': [0, 1, 2], 'explain': '.选class、#选id、空格选后代；选择器既能选标签也能选 class/id。'},
          {'type': 'open', 'question': '你在开发者工具里对一个元素点"复制 selector"得到 `#list > li.item`，你怎么在 bs4 里用它抠出所有 li 的文字？写 2 行。', 'answer': 'items = soup.select("#list > li.item"); [print(i.text) for i in items]。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['SELECT', 'SELECTOR', 'SELECTONE']},
      ],
      tasks=[
          '在开发者工具对一个元素"复制 selector"，粘进 soup.select() 验证能抠到。',
          '用 `.price` 选择器抠你目标页面的价格，和之前 find_all 的结果对比是否一致。',
          '写一个 `div#content p` 的选择器，抠出正文段落并打印数量。',
      ]),

    L('r4l4', '提取文本、属性与层级遍历', '🧭',
      r"""## 从"找到一个标签"到"拿到干净数据"

真实页面脏得很：文字带换行、前后空格、嵌套标签。三板斧收拾它：

### 1. 取文字并清洗
```python
title = soup.select_one('.title').get_text(strip=True)   # strip 去首尾空白
```

### 2. 取属性（链接/图片地址）
```python
img = soup.select_one('img')
print(img.get('src'))        # 图片地址
print(img.get('alt', ''))    # 取 alt，没有就给空串
```

### 3. 层级遍历
```python
# 父、子、兄弟
tag.parent          # 父节点
tag.children        # 直接子节点（迭代器）
tag.find_next_sibling()  # 下一个兄弟
# 在容器内逐个抠子项
for li in soup.select('ul.news li'):
    print(li.get_text(strip=True))
```

### 实战套路（记住这个模板）
> **先定位容器 → 再在里面逐个抠子项 → 每条存成字典 → 字典装进列表 → 存文件。**

这就是一整页数据变 Excel 的标准流水线。""",
      takeaway=r""".get_text(strip=True) 取干净文字；.get('src') 取图片/链接地址；parent/children/sibling 做层级遍历。标准流水线：**定位容器→逐个抠子项→每条存字典→装列表→存文件**。""",
      words=[
          {'en': 'GETTEXT', 'zh': '.get_text(strip=True)：取文字并去掉首尾空白'},
          {'en': 'ATTRIBUTE', 'zh': '.get("src")：取标签属性（如图片/链接地址）'},
          {'en': 'TRAVERSE', 'zh': '遍历：parent 父 / children 子 / sibling 兄弟'},
      ],
      exercises=[
          {'type': 'choice', 'question': '`get_text(strip=True)` 的 strip=True 作用是？', 'options': ['去掉文字首尾空白', '只取英文', '转大写'], 'answer': 0, 'explain': 'strip 去除首尾空白与换行，拿到干净文字。'},
          {'type': 'choice', 'question': '取一个 img 标签的图片地址，用？', 'options': [".get('src')", '.text', '.href'], 'answer': 0, 'explain': '图片地址在 src 属性，用 .get("src") 取。'},
          {'type': 'fill', 'question': '取标签的父节点用 ______ 属性（填 parent/child/sibling 之一）。', 'answer': 'parent', 'explain': 'tag.parent 是父节点。'},
          {'type': 'tap', 'question': '关于层级遍历，说法对的有？（多选）', 'options': ['.parent 取父', '.children 取直接子', 'find_next_sibling 取下个兄弟', '只能从根往下不能反向'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'bs4 支持 parent/children/sibling 双向导航；并非只能从根往下。'},
          {'type': 'open', 'question': '你抠出一条新闻，文字里混着 <span> 和 <br>，怎么拿到"纯文字且没首尾空格"的结果？写 2 行。', 'answer': 'node = soup.select_one(".news"); text = node.get_text(strip=True)。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['GETTEXT', 'ATTRIBUTE', 'TRAVERSE']},
      ],
      tasks=[
          '用 get_text(strip=True) 抠一个标题，对比加不加 strip 的差异。',
          '抠一个页面里所有 img 的 src，打印出来（注意相对路径要拼绝对 URL）。',
          '对一个 <ul><li> 列表，用"定位容器→逐个抠 li"的模板把内容存成字典列表。',
      ]),
])

# ============================ 第5章 正则与 JSON ============================
ch5 = CH('正则与 JSON', [
    L('r5l1', '正则表达式：用模式捞文字', '🔍',
      r"""## 当 HTML 结构太乱，正则是备用刀

BeautifulSoup 是主力，但有些数据藏在**不规则文本/JS 变量**里，正则反而快。正则 = 用**模式**匹配字符串。

### 最常用的几个符号
| 符号 | 含义 | 例子 |
|---|---|---|
| `\d` | 一个数字 | `\d+` 匹配一串数字 |
| `\w` | 字母数字下划线 | |
| `.` | 任意字符(除换行) | |
| `*` `+` | 重复 0+ / 1+ 次 | |
| `[]` | 字符集合 | `[a-z]` |
| `()` | 分组(捕获) | |

### 基础用法
```python
import re
text = '价格 39 元，库存 12 件'
nums = re.findall(r'\d+', text)     # ['39', '12']
print(nums)
```

### 心态
正则上手有点劝退，但**抓数字、邮箱、日期、URL 这种"模式固定"的东西它贼好用**。bs4 搞不定的边角，它来补刀。第 2 节讲分组提取。""",
      figures=[{'key': 'regex_match', 'caption': '🔍 正则像"模式滤网"：\\d+ 从乱文本里只捞出数字串；适合邮箱/日期/URL 这类固定模式'}],
      takeaway=r"""正则是用"模式"捞文字的备用刀：\d 数字、\w 词字符、+、* 重复、() 分组、[] 集合。抓数字/邮箱/日期/URL 这种固定模式它贼好用。bs4 搞不定的边角它补刀。""",
      words=[
          {'en': 'REGEX', 'zh': '正则(re)：用模式串匹配/提取文本的工具'},
          {'en': 'PATTERN', 'zh': '模式：正则里描述"要匹配啥"的表达式'},
          {'en': 'FINDALL', 'zh': 're.findall：找出所有匹配，返回列表'},
      ],
      exercises=[
          {'type': 'choice', 'question': '正则里 `\\d+` 表示？', 'options': ['一串数字', '一个字母', '任意字符'], 'answer': 0, 'explain': '\\d 数字，+ 一个或多个，合起来匹配连续数字。'},
          {'type': 'choice', 'question': '`re.findall(r"\\d+", "a1b22c333")` 返回？', 'options': ['["1","22","333"]', '["123"]', '[]'], 'answer': 0, 'explain': 'findall 找所有连续数字串，分别 1/22/333。'},
          {'type': 'fill', 'question': '正则里表示"重复一次或多次"的量词符号是 ___ 。', 'answer': '+', 'explain': '+ 表示前面元素出现 1 次或多次；* 是 0 次或多次。'},
          {'type': 'tap', 'question': '关于正则，说法对的有？（多选）', 'options': ['\\d 匹配数字', '() 用于分组捕获', 'findall 返回所有匹配', '正则可替代 bs4 解析一切 HTML'], 'multi': True, 'answer': [0, 1, 2], 'explain': '\\d/()/findall 都对；但"正则替代 bs4 解析一切 HTML"不现实，复杂结构还是 bs4 稳。'},
          {'type': 'open', 'question': '什么场景下你会优先用正则而不是 bs4？举一个具体例子。写 2 句。', 'answer': '当数据嵌在 JS 变量或纯文本里（如 "var data={"abc":123}" 或一段日志），结构乱无标签时，用正则按模式提取比 bs4 更直截了当。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['REGEX', 'PATTERN', 'FINDALL']},
      ],
      tasks=[
          '用 re.findall(r"\\d+", 一段含多个数字的文本) 把所有数字抠出来。',
          '写一个正则匹配常见邮箱（含 @ 和域名），从一段文本里抠出所有邮箱。',
          '从一段"价格：39 元 销量：1200"的文字里，分别用正则抠出价格和销量数字。',
      ]),

    L('r5l2', '分组与替换：提取结构化字段', '📦',
      r"""## 不止"找到"，还要"拆开取"

`()` 在正则里是**分组**，能让你把匹配里的某一段单独拎出来。

### 用 group 取分组
```python
import re
s = '用户: 小明, 年龄: 14'
m = re.search(r'用户:\s*(\w+),\s*年龄:\s*(\d+)', s)
if m:
    print(m.group(1))   # 小明
    print(m.group(2))   # 14
```

### 用 findall + 分组
```python
text = 'A:10 B:20 C:30'
pairs = re.findall(r'(\w+):(\d+)', text)   # [('A','10'),('B','20'),('C','30')]
```

### 替换：re.sub
```python
clean = re.sub(r'\s+', ' ', '太   多    空格')   # 多个空格压成一个
```

### 何时用
当你要的是"**字段名:值**""**键=值**"这种成对结构，分组 + findall 一键变元组列表，再转字典，舒服。""",
      takeaway=r"""() 分组能把匹配里的某段单独取出：re.search 用 group(1)(2) 取，re.findall 带分组直接返回元组列表。re.sub 做替换。抓"键=值/名:值"成对结构时，分组最香。""",
      words=[
          {'en': 'GROUP', 'zh': '分组：正则里 () 圈出要单独提取的片段'},
          {'en': 'SUB', 'zh': 're.sub：按正则替换文本，如压空格'},
          {'en': 'SEARCH', 'zh': 're.search：找第一个匹配，返回可取分组的对象'},
      ],
      exercises=[
          {'type': 'choice', 'question': '`re.search(r"年龄:(\\d+)", s).group(1)` 取到的是？', 'options': ['括号里捕获的年龄数字', '整个匹配串', '报错'], 'answer': 0, 'explain': 'group(1) 取第一个分组，即年龄数字。'},
          {'type': 'choice', 'question': '`re.sub(r"\\s+", " ", t)` 的作用是？', 'options': ['多个空白压成一个空格', '删除所有空格', '在空格前加字'], 'answer': 0, 'explain': 'sub 把连续空白替换为单个空格。'},
          {'type': 'fill', 'question': '要在正则里"圈出要单独提取的片段"，应使用的符号是 ___（填一对括号类型）。', 'answer': '()', 'explain': '圆括号 () 定义捕获分组。'},
          {'type': 'tap', 'question': '关于分组与替换，说法对的有？（多选）', 'options': ['() 定义分组', 'group(1) 取第一组', 're.sub 做替换', 'findall 带分组返回元组列表'], 'multi': True, 'answer': [0, 1, 2, 3], 'explain': '四条全对：()分组、group取组、sub替换、findall带分组返回元组。'},
          {'type': 'open', 'question': '一段文本是 "name=小明;age=14;city=成都"，你怎么用正则一次性拆成字典？写 2-3 行思路。', 'answer': 'pairs = re.findall(r"(\\w+)=([^;]+)", text); d = dict(pairs)。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['GROUP', 'SUB', 'SEARCH']},
      ],
      tasks=[
          '用 re.search + 分组，从 "价格:39 库存:12" 里分别取出价格和库存数字。',
          '用 re.findall(r"(\\w+):(\\d+)", ...) 把"名:值"对变成元组列表再转字典。',
          '用 re.sub 把一段多余空白的文本压成单词间单个空格。',
      ]),

    L('r5l3', 'JSON 与接口：直接吃现成的结构化数据', '🍱',
      r"""## 最爽的抓取：人家把数据都摆好了

很多网站的数据通过**接口（API）**返回 **JSON**——一种天然结构化的文本（键值对、列表），解析零成本。

### JSON 长这样
```json
{
  "title": "Python 入门",
  "price": 39,
  "tags": ["入门", "编程"]
}
```

### Python 处理 JSON
```python
import requests, json

r = requests.get(api_url, headers=headers)
data = r.json()          # 直接解析成 字典/列表
print(data['title'])
for t in data['tags']:
    print(t)
```
`r.json()` 把 JSON 文本变成 Python 的 dict/list，**后面存取和前面学的字典完全一样**。

### 为什么爽
- 不用 bs4 抠 HTML，没有标签干扰
- 字段清晰，直接按名字取
- 常配合 `params` 翻页、筛选

> 找接口的方法回看第 2 章：F12 → Network → 找返回 JSON 的那条 XHR/Fetch。""",
      figures=[{'key': 'api_json', 'caption': '🍱 接口直接返回 JSON：requests.get→r.json() 变 dict/list，按名字取字段，比抠 HTML 省事'}],
      takeaway=r"""接口返回的 JSON 用 `r.json()` 一键变 dict/list，存取和字典一模一样——**免抠 HTML、字段清晰、还能拼 params 翻页**，是爬虫最爽的姿势。找接口回看 F12 Network。""",
      words=[
          {'en': 'JSON', 'zh': 'JSON：一种键值对/列表结构的数据格式，爬虫最爱'},
          {'en': 'API', 'zh': '接口：网站用来返回结构化数据（常是 JSON）的网址'},
          {'en': 'RJASON', 'zh': 'r.json()：把响应 JSON 文本解析成 Python 对象'},
      ],
      exercises=[
          {'type': 'choice', 'question': '`r.json()` 把 JSON 响应变成？', 'options': ['Python 的 dict/list', '字符串', 'bytes'], 'answer': 0, 'explain': 'json() 解析成原生 dict/list，后续按名存取。'},
          {'type': 'choice', 'question': '相比抓 HTML 再解析，直接打接口拿 JSON 的好处是？', 'options': ['字段清晰、免抠标签', '更慢', '必须登录'], 'answer': 0, 'explain': 'JSON 结构化、无标签干扰，直接按名取，省事。'},
          {'type': 'fill', 'question': '用 requests 拿到 JSON 后，调用 r.______() 即可解析成对象。', 'answer': 'json', 'explain': 'r.json() 解析响应为 dict/list。'},
          {'type': 'tap', 'question': '关于 JSON 与接口，说法对的有？（多选）', 'options': ['JSON 是键值对/列表结构', 'r.json() 得到 dict/list', '接口常在 Network 的 XHR/Fetch 里', 'JSON 一定比 HTML 难解析'], 'multi': True, 'answer': [0, 1, 2], 'explain': '前三条对；JSON 其实比 HTML 好解析，不是更难。'},
          {'type': 'open', 'question': '你发现目标数据在接口返回的 JSON 里是个列表，每项是个含 title/url 的对象。写 3 行核心代码把全部 title 打印出来。', 'answer': 'data = r.json(); items = data["list"]; [print(i["title"]) for i in items]。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['JSON', 'API', 'RJASON']},
      ],
      tasks=[
          '找一个返回 JSON 的接口（第2章方法），用 r.json() 解析并打印某个字段。',
          '把一个 JSON 响应里的列表项逐条存成字典列表，再打印长度。',
          '对比"抓同一站点的 HTML 版"和"抓它的 JSON 接口版"，体会哪种更省事。',
      ]),
])

# ============================ 第6章 数据存储 ============================
ch6 = CH('数据存储', [
    L('r6l1', 'CSV：表格界的老大哥', '📊',
      r"""## 爬回来不存，等于白爬

CSV（逗号分隔）是**最通用的表格格式**，Excel、数据库、pandas 都能直接吃。爬虫落地首选。

### 用 csv 模块写
```python
import csv

rows = [
    {'title': 'A', 'price': 30},
    {'title': 'B', 'price': 25},
]
with open('books.csv', 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=['title', 'price'])
    w.writeheader()
    w.writerows(rows)
```
> `utf-8-sig` 关键：让 Excel 打开中文不乱码（带 BOM）。`newline=''` 防止 Windows 下多出空行。

### 读回来
```python
with open('books.csv', encoding='utf-8-sig', newline='') as f:
    for row in csv.DictReader(f):
        print(row['title'], row['price'])
```

### 何时用 CSV
数据是一条条**同字段**的记录（榜单、商品、新闻），CSV 最轻量。嵌套结构多的再用 JSON（下一节）。""",
      takeaway=r"""爬回数据用 csv.DictWriter 落地成 .csv，Excel 直接开。**utf-8-sig 防 Excel 中文乱码、newline='' 防空行**是两条保命细节。同字段记录用 CSV 最轻量。""",
      words=[
          {'en': 'CSV', 'zh': 'CSV：逗号分隔的表格文本，Excel/数据库通用'},
          {'en': 'DICTWRITER', 'zh': 'DictWriter：按字段名把字典列表写成 CSV'},
          {'en': 'BOM', 'zh': 'utf-8-sig：带 BOM 的 UTF-8，让 Excel 正确识别中文'},
      ],
      code=r"""rows = [
    {'title': 'A', 'price': 30},
    {'title': 'B', 'price': 25},
    {'title': 'C', 'price': 40},
]
# 浏览器迷你 Python 不能真写硬盘，这里模拟写出内容
header = ','.join(rows[0].keys())
lines = [header]
for r in rows:
    lines.append(','.join(str(r[k]) for k in r))
print('\n'.join(lines))""",
      exercises=[
          {'type': 'choice', 'question': '为什么写 CSV 给 Excel 用推荐 utf-8-sig 而不是 utf-8？', 'options': ['带 BOM，Excel 中文不乱码', '更快', '文件更小'], 'answer': 0, 'explain': 'utf-8-sig 带 BOM，Excel 能正确识别中文编码。'},
          {'type': 'choice', 'question': '`csv.DictWriter` 适合写什么形状的数据？', 'options': ['字典的列表（每条同字段）', '单个数字', '一张图片'], 'answer': 0, 'explain': 'DictWriter 按字段名写字典列表，正好对应表格。'},
          {'type': 'fill', 'question': '写 CSV 时加 newline='' 是为了防止在 ______ 系统下出现多余空行。', 'answer': 'windows', 'explain': 'Windows 下不设这个参数容易每行多空行。'},
          {'type': 'coding', 'question': '把下面的字典列表用逗号拼成一个 CSV 表头+两行（不用真写文件），目标输出含 title,price 和 A,30。', 'starter': "rows = [{'title':'A','price':30},{'title':'B','price':25}]\n# 拼接成 CSV 文本并打印", 'expect': 'A,30', 'hint': "用 ','.join(...) 拼每行。"},
          {'type': 'tap', 'question': '关于 CSV 存储，说法对的有？（多选）', 'options': ['适合同字段的记录', 'Excel 能直接打开', 'utf-8-sig 防乱码', '适合任意深度嵌套'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'CSV 适合扁平同字段、Excel 友好、utf-8-sig 防乱码；深度嵌套不适合 CSV。'},
          {'type': 'open', 'question': '你爬了 500 条商品（名称/价格/销量），为什么 CSV 比直接堆进一个 .txt 更合适？写 2-3 句。', 'answer': 'CSV 字段规整、Excel/pandas 可直接分析筛选，而 txt 自由格式难以结构化检索；500 条用表格最方便。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['CSV', 'DICTWRITER', 'BOM']},
      ],
      tasks=[
          '把第 4 章抠到的"字典列表"用 csv.DictWriter 写成 books.csv，用 Excel 打开确认中文正常。',
          '用 csv.DictReader 把刚写的文件读回来，打印每一行。',
          '故意用普通 utf-8 写一份，再用 Excel 打开对比乱码差异，体会 BOM 的作用。',
      ]),

    L('r6l2', 'JSON 文件：嵌套数据随便存', '🗄️',
      r"""## 当数据有"层次"，JSON 比 CSV 能装

CSV 是扁平的（一行一记录）。但爬虫常遇到**嵌套**：一条记录里还有列表、子对象。这种用 **JSON 文件**最自然。

### 写 JSON
```python
import json

data = {
    'keyword': 'python',
    'results': [
        {'title': 'A', 'tags': ['入门', '编程']},
        {'title': 'B', 'tags': ['进阶']},
    ]
}
with open('out.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
```
> `ensure_ascii=False` 必加——否则中文变成 `\uXXXX` 鬼画符。`indent=2` 让文件好看。

### 读 JSON
```python
with open('out.json', encoding='utf-8') as f:
    data = json.load(f)
print(data['results'][0]['title'])
```

### CSV vs JSON 怎么选
- **扁平同字段** → CSV（给人看、给 Excel）
- **有嵌套/层级** → JSON（给程序继续处理）""",
      takeaway=r"""嵌套数据用 json.dump 存 .json，`ensure_ascii=False` 保中文、`indent=2` 好看。扁平同字段选 CSV，有层级嵌套选 JSON——这是落地时的二选一。""",
      words=[
          {'en': 'JSON', 'zh': 'JSON：可表达嵌套（对象/数组）的数据格式'},
          {'en': 'DUMP', 'zh': 'json.dump：把 Python 对象写入 JSON 文件'},
          {'en': 'ENSUREASCII', 'zh': 'ensure_ascii=False：让中文正常写入而非转义'},
      ],
      code=r"""data = {
    'keyword': 'python',
    'results': [
        {'title': 'A', 'tags': ['入门', '编程']},
        {'title': 'B', 'tags': ['进阶']},
    ],
}
# 真实项目用 json.dumps(data, ensure_ascii=False, indent=2) 存盘；
# 沙盒里咱们用纯 Python 看一眼这份"嵌套结构"长什么样：
print('关键词:', data['keyword'])
for i, r in enumerate(data['results'], 1):
    print(i, r['title'], '标签:', '、'.join(r['tags']))""",
      exercises=[
          {'type': 'choice', 'question': '写 JSON 时 `ensure_ascii=False` 的作用是？', 'options': ['让中文正常写入而非变 \\uXXXX', '压缩文件', '加密'], 'answer': 0, 'explain': '默认 ensure_ascii=True 会把中文转成 \\u 转义，设 False 才保留中文。'},
          {'type': 'choice', 'question': '一条记录里含"标签列表"这种嵌套，更适合存成？', 'options': ['JSON', 'CSV', '纯文本'], 'answer': 0, 'explain': '嵌套结构 JSON 原生支持，CSV 扁平难表达。'},
          {'type': 'fill', 'question': '从 JSON 文件读回对象用 json.______(f)（填 load 或 dump）。', 'answer': 'load', 'explain': 'json.load 从文件读，json.dump 写入文件。'},
          {'type': 'tap', 'question': '关于 JSON 存储，说法对的有？（多选）', 'options': ['支持嵌套结构', 'ensure_ascii=False 保中文', 'indent 让文件美观', '只能存扁平数据'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'JSON 支持嵌套、能保中文、可缩进；并非只能扁平。'},
          {'type': 'open', 'question': '你抓的每条新闻含"标题、作者、多条评论(列表)"。为什么 JSON 比 CSV 更适合存它？写 2-3 句。', 'answer': '评论是列表嵌套在新闻对象里，JSON 能直接表达这种层级；CSV 扁平表难以自然容纳变长评论列表。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['JSON', 'DUMP', 'ENSUREASCII']},
      ],
      tasks=[
          '把第 5 章接口拿到的嵌套数据用 json.dump(ensure_ascii=False, indent=2) 存成 out.json。',
          '用 json.load 读回来，打印其中某个嵌套字段（如 results[0]["title"]）。',
          '对比同一份数据存 CSV 和 JSON 的差异，想清楚什么时候用哪个。',
      ]),

    L('r6l3', 'Excel：用 openpyxl 输出 .xlsx', '📈',
      r"""## 要交给非程序员？直接给 .xlsx

CSV 虽好，但有些人就认 Excel 的 `.xlsx`。用 `openpyxl` 能直接生成真·Excel 文件。

### 写 xlsx
```python
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(['标题', '价格'])          # 表头
ws.append(['A', 30])
ws.append(['B', 25])
wb.save('books.xlsx')
```

### 读 xlsx
```python
from openpyxl import load_workbook
wb = load_workbook('books.xlsx')
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
```

### 注意
- `openpyxl` 要 `pip install openpyxl`
- 大批量数据（几万行）CSV 更快；xlsx 适合"给人看的成品报表"
- 和 CSV 一样，中文无需额外 BOM（xlsx 本身就是 UTF-8 容器）

> 一条流水线打通：**requests 抓 → bs4/json 解析 → 整理成字典列表 → openpyxl 存 xlsx**，这就是"学完能交差"的闭环。""",
      figures=[{'key': 'data_store', 'caption': '💾 落地三选：CSV(扁平/Excel友好) · JSON(嵌套/程序友好) · xlsx(给人看的成品报表)'}],
      takeaway=r"""给非程序员交付用 openpyxl 存 .xlsx：`wb=Workbook(); ws.append(...); wb.save()`。CSV 快、JSON 装嵌套、xlsx 给人看——三条落地路线打通，爬虫才真正"能用"。""",
      words=[
          {'en': 'OPENPYXL', 'zh': 'openpyxl：读写 .xlsx Excel 文件的库'},
          {'en': 'WORKBOOK', 'zh': 'Workbook：一个 Excel 工作簿对象'},
          {'en': 'XLSX', 'zh': 'xlsx：Excel 的现代文件格式，程序可直接生成'},
      ],
      exercises=[
          {'type': 'choice', 'question': '要生成给人看的 .xlsx，用哪个库？', 'options': ['openpyxl', 'csv', 'json'], 'answer': 0, 'explain': 'openpyxl 专门读写 xlsx。'},
          {'type': 'choice', 'question': 'openpyxl 写表头用？', 'options': ['ws.append(["标题","价格"])', 'ws.write("标题")', 'wb.header(...)'], 'answer': 0, 'explain': 'ws.append 按行追加，先 append 表头再 append 数据行。'},
          {'type': 'fill', 'question': 'openpyxl 保存文件用 wb.______("books.xlsx")（填 save/load）。', 'answer': 'save', 'explain': 'wb.save 写盘；load_workbook 才是读。'},
          {'type': 'tap', 'question': '关于三种落地格式，对应对的有的？（多选）', 'options': ['CSV 快且 Excel 友好', 'JSON 适合嵌套', 'xlsx 给人看的成品', '三者完全一样没区别'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'CSV 轻快、JSON 装嵌套、xlsx 交人；三者定位不同。'},
          {'type': 'open', 'question': '老板要你"每周一发一份能直接筛选的 Excel"。你的爬虫最后一步该用什么、注意什么？写 2-3 句。', 'answer': '用 openpyxl 生成 xlsx，第一行放表头字段，数据按行 append，最后 wb.save；注意字段命名清晰、类型正确便于筛选。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['OPENPYXL', 'WORKBOOK', 'XLSX']},
      ],
      tasks=[
          '把前面攒的字典列表用 openpyxl 写成 books.xlsx（含表头）。',
          '用 load_workbook 读回来，iter_rows 打印每一行。',
          '思考并记下：你最终要交付给谁？据此选 CSV / JSON / xlsx 其中一种作为默认落地格式。',
      ]),
])

# ============================ 第7章 进阶技巧 ============================
ch7 = CH('进阶技巧', [
    L('r7l1', '分页爬取：把"下一页"摁到底', '📖',
      r"""## 数据不止一页，那就一页页摁

绝大多数列表都是分页的。两种常见套路：

### 套路 A：URL 参数翻页
页码体现在 URL（第 2 章学过）：
```python
for page in range(1, 11):
    url = f'https://x.com/list?page={page}'
    r = requests.get(url, headers=headers, timeout=10)
    parse(r.text)
    sleep(1)     # 礼貌：翻页也带延迟
```

### 套路 B：接口翻页（offset / cursor）
接口常用 `offset=0&limit=20` 或"下一页游标"：
```python
offset = 0
while True:
    data = requests.get(api, params={'offset': offset, 'limit': 20}).json()
    if not data['items']:
        break
    save(data['items'])
    offset += 20
```

### 关键细节
- **一定要有终止条件**（页码上限 / 返回空 / 出现重复），否则死循环
- **翻页也带延迟**，别把人家的分页接口打爆
- 记录"爬到哪一页"，断了能从中间续""",
      takeaway=r"""分页两套路：URL 参数翻页（page=）和接口 offset/cursor 翻页。**必须有终止条件（上限/空返回/重复）防死循环，翻页也带延迟**，并记下进度以便断点续爬。""",
      words=[
          {'en': 'PAGINATION', 'zh': '分页：数据分成多页，逐页抓取'},
          {'en': 'OFFSET', 'zh': 'offset：接口里的"从第几条开始取"参数'},
          {'en': 'CURSOR', 'zh': '游标：接口返回"下一页指针"，用它接着翻'},
      ],
      exercises=[
          {'type': 'choice', 'question': '分页爬取最重要的安全保障是？', 'options': ['有终止条件防死循环', '越快越好', '忽略重复'], 'answer': 0, 'explain': '必须有上限/空返回/重复检测等终止条件，否则死循环。'},
          {'type': 'choice', 'question': '接口翻页常用哪两个参数思路？', 'options': ['offset/limit 或 cursor 游标', '只改 URL 路径', '完全随机'], 'answer': 0, 'explain': 'offset+limit 或下一页 cursor 是接口分页主流。'},
          {'type': 'fill', 'question': '翻页时加 sleep(1) 是为了保持______（填"礼貌"或"速度"）。', 'answer': '礼貌', 'explain': '延迟降低请求频率，避免打爆服务器，也是反爬友好。'},
          {'type': 'tap', 'question': '关于分页，说法对的有？（多选）', 'options': ['URL 参数可翻页', '接口可用 offset', '要防死循环', '翻页不用延迟'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'URL 参数、接口 offset 都能翻页且要防死循环；翻页也应加延迟。'},
          {'type': 'open', 'question': '你爬一个页，发现第 11 页内容和第 1 页一样（网站兜底返回首页）。你的循环怎么避免无限重复？写 2-3 句。', 'answer': '爬前记录已见URL/内容哈希，若新页与上一页重复或已出现过则 break；或设最大页数上限。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['PAGINATION', 'OFFSET', 'CURSOR']},
      ],
      tasks=[
          '找一个带 ?page= 的列表站，写循环爬前 5 页，每页 sleep(1)。',
          '若目标是接口分页，用 offset 循环抓取直到返回空列表。',
          '给你的爬虫加一个"已爬页码记录"，模拟中断后从断点继续。',
      ]),

    L('r7l2', '会话与登录：带着 Cookie 进门', '🔐',
      r"""## 有些数据，得"登录后"才给看

会员内容、个人中心，必须带着**登录态（Cookie）**才能访问。第 3 章提过 `Session()`，这里实战。

### 模拟登录流程
```python
s = requests.Session()
# 1) 先 GET 登录页，拿可能需要的 token（有些站点有）
# 2) POST 账号密码
login = s.post('https://x.com/login',
               data={'user': 'me', 'pwd': '123'},
               headers=headers)
# 3) 之后用同一个 s 访问，Cookie 自动带上
r = s.get('https://x.com/vip-data')
```

### 两个坑
- **验证码 / 滑块**：很多登录有防护，纯 requests 过不了——这种要么手动拿 Cookie 粘进 headers，要么上 Selenium（下一节）
- **Cookie 会过期**：别把 Cookie 写死缓存太久，失效了重新登录

### 心态
能模拟登录是能力，但**只用在你自己有权限的账号/数据上**。拿别人的、批量撞库，那是违法的边缘。""",
      takeaway=r"""带登录态用 requests.Session()：POST 登录一次，后续同 s 请求自动带 Cookie。**验证码/滑块可能过不了，可考虑手动粘 Cookie 或 Selenium**；且只用于自己有权限的账号，别越界。""",
      words=[
          {'en': 'COOKIE', 'zh': 'Cookie：服务器发的"身份凭证"，带着它算登录态'},
          {'en': 'LOGIN', 'zh': '登录：拿到并维持 Cookie 的过程'},
          {'en': 'CAPTCHA', 'zh': '验证码：防止自动登录的防护，requests 常过不了'},
      ],
      exercises=[
          {'type': 'choice', 'question': '用 Session 登录后访问其他页，Cookie 怎么处理？', 'options': ['自动带上，不用管', '每次手动复制', '登录失效也不影响'], 'answer': 0, 'explain': 'Session 自动维护 Cookie，后续请求自动带。'},
          {'type': 'choice', 'question': '遇到登录验证码/滑块，requests 通常？', 'options': ['过不了，需手动 Cookie 或 Selenium', '轻松绕过', '自动消失'], 'answer': 0, 'explain': '这类防护专防自动登录，纯 requests 难破，需换方案。'},
          {'type': 'fill', 'question': '服务器发的"身份凭证"叫 ______（填 Cookie 或 Token 之一，本课指前者）。', 'answer': 'Cookie', 'explain': 'Cookie 是会话身份凭证，维持登录态。'},
          {'type': 'tap', 'question': '关于登录爬取，说法对的有？（多选）', 'options': ['Session 自动带 Cookie', '验证码可能需换方案', '只用于自己有权限的账号', 'Cookie 永不过期'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'Session 带 Cookie、验证码需换方案、仅用于自己账号；Cookie 会过期不是永久。'},
          {'type': 'open', 'question': '你"借用"朋友的账号密码写爬虫爬他的私信。这有什么问题和风险？写 2-3 句。', 'answer': '未经授权访问他人私信涉嫌侵犯隐私甚至违法；且撞库/盗用账号是犯罪边缘，绝不可为。仅用自己有权限的账号。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['COOKIE', 'LOGIN', 'CAPTCHA']},
      ],
      tasks=[
          '用一个你自己的测试账号，Session POST 登录某站点，再访问需登录的页面验证成功。',
          '如果登录有验证码，尝试手动从浏览器复制 Cookie 粘进 headers 访问一次。',
          '把"登录态会过期"写进你的代码注释，设计"失效就重新登录"的逻辑草图。',
      ]),

    L('r7l3', '代理与随机延迟：别把人家薅秃', '🛡️',
      r"""## 爬得猛，容易被"拉黑"

当你请求频率过高，目标可能**封你 IP** 或弹验证码。两个缓解手段：

### 随机延迟
```python
from time import sleep
import random
sleep(random.uniform(1, 3))    # 每次随机睡 1~3 秒
```
比固定 sleep(2) 更"像人"——真人不会精确到秒。

### 代理 IP
```python
proxies = {'http': 'http://1.2.3.4:8080', 'https': 'http://1.2.3.4:8080'}
r = requests.get(url, proxies=proxies, timeout=10)
```
代理让你"换张脸"再请求，避免单个 IP 被盯死。

### 心态（很重要）
这些是**自我保护 + 降低对人家压力**的手段，不是让你更狠地薅。配合第 8 章的礼貌与合规，**做个有分寸的爬虫**。滥用代理做坏事，IP 池、账号都可能牵连法律责任。""",
      figures=[{'key': 'proxy_rotate', 'caption': '🛡️ 随机延迟+代理轮换：降低单IP压力、像真人节奏；但目的是礼貌而非更狠地薅'}],
      takeaway=r"""随机延迟(random.uniform)比固定 sleep 更像人；代理(proxies)让你换 IP 避免被封。但这些都是**降压力、自我保护**，不是让你更狠薅——配合合规做个有分寸的爬虫。""",
      words=[
          {'en': 'PROXY', 'zh': '代理：中转服务器，让你换 IP 发起请求'},
          {'en': 'DELAY', 'zh': '延迟：请求间隔，降低频率避免被封'},
          {'en': 'RATE', 'zh': '频率/速率：单位时间请求数，越低越礼貌'},
      ],
      exercises=[
          {'type': 'choice', 'question': '`random.uniform(1, 3)` 相比固定 `sleep(2)` 的好处是？', 'options': ['请求间隔更像真人', '更快', '更慢'], 'answer': 0, 'explain': '随机间隔打破规律，更接近人工访问节奏。'},
          {'type': 'choice', 'question': 'requests 使用代理要传哪个参数？', 'options': ['proxies', 'proxy', 'agent'], 'answer': 0, 'explain': 'requests.get(url, proxies={...}) 设置代理。'},
          {'type': 'fill', 'question': '代理和延迟的目的是降低对目标服务器的______（填"压力"或"速度"）。', 'answer': '压力', 'explain': '本质是降频降压、自我保护并减少干扰。'},
          {'type': 'tap', 'question': '关于代理与延迟，说法对的有？（多选）', 'options': ['随机延迟更像人', 'proxies 换 IP', '用于降压力', '用来更狠地薅数据'], 'multi': True, 'answer': [0, 1, 2], 'explain': '随机延迟、proxies 换 IP、降压力都对；"更狠薅"违背合规本意。'},
          {'type': 'open', 'question': '你爬一个站，刚跑 100 页 IP 就被封了。你下一步该怎么调整才合理？写 2-3 句。', 'answer': '先降速加随机延迟、减少并发；必要时加代理轮换 IP；同时检查是否违反 robots/频率限制，必要时暂停或联系授权。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['PROXY', 'DELAY', 'RATE']},
      ],
      tasks=[
          '把你的爬虫请求间隔从固定 sleep 改成 random.uniform(1,3)。',
          '（若有条件）配置一个代理，验证 requests 能经它发出请求。',
          '估算你目标站点"多久抓一次算礼貌"，写进代码注释作为频率上限。',
      ]),

    L('r7l4', '动态页面与 Selenium：对付 JS 渲染', '🎬',
      r"""## 有些页面，requests 看不见内容

前面说过：很多数据由 **JS 在浏览器里动态生成**，初始 HTML 是空的，`requests` 抓回来啥也没有。这种叫**动态页面**。

### 解法：让"真浏览器"替你跑
`Selenium` 能驱动 Chrome 真实渲染页面，等 JS 跑完再拿 HTML：
```python
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()
driver.get(url)
driver.implicitly_wait(5)            # 等元素出现
titles = driver.find_elements(By.CSS_SELECTOR, '.title')
for t in titles:
    print(t.text)
driver.quit()
```

### 代价
- 要装 **浏览器驱动**（ChromeDriver），比 requests 重
- 慢、占资源，**能不用就不用**——优先确认数据是否藏在接口 JSON（第 5 章），能打接口就别开浏览器

### 决策顺序（记牢）
> 先看 Network 有没有 JSON 接口 → 有就 requests 打接口；**没有才**上 Selenium 渲染。

> 注：Selenium 需在**你自己电脑**装浏览器和驱动，浏览器迷你 Python 跑不了，这里只讲思路，动手在你本机。""",
      figures=[{'key': 'selenium_dynamic', 'caption': '🎬 动态页面：初始HTML为空，JS渲染后才出数据；Selenium驱动真浏览器等渲染完再抓'}],
      takeaway=r"""动态页面(JS 渲染、初始 HTML 空)用 Selenium 驱动真浏览器等渲染完再抓。**但优先确认有没有 JSON 接口可打，能 requests 就别开浏览器**——Selenium 重、慢、要装驱动。""",
      words=[
          {'en': 'SELENIUM', 'zh': 'Selenium：驱动真实浏览器自动化渲染/操作的库'},
          {'en': 'DYNAMIC', 'zh': '动态页面：内容由 JS 渲染，初始 HTML 拿不到'},
          {'en': 'DRIVER', 'zh': '驱动：如 ChromeDriver，让 Selenium 控制浏览器'},
      ],
      exercises=[
          {'type': 'choice', 'question': 'requests 抓回来是空 HTML，最可能是？', 'options': ['内容是 JS 动态渲染的', '网络断了', '编码错了'], 'answer': 0, 'explain': '动态页面初始 HTML 为空，数据靠 JS 渲染，requests 看不到。'},
          {'type': 'choice', 'question': '对付动态页面，优先应该？', 'options': ['先看有没有 JSON 接口可打', '立刻上 Selenium', '放弃'], 'answer': 0, 'explain': '先看 Network 是否有接口 JSON，能打接口就别开重型的 Selenium。'},
          {'type': 'fill', 'question': 'Selenium 控制浏览器需要安装对应的______（如 ChromeDriver）。', 'answer': '驱动', 'explain': '驱动是 Selenium 与浏览器之间的桥梁。'},
          {'type': 'tap', 'question': '关于 Selenium，说法对的有？（多选）', 'options': ['能渲染 JS 动态页面', '比 requests 重且慢', '优先于打接口使用', '需装浏览器驱动'], 'multi': True, 'answer': [0, 1, 3], 'explain': 'Selenium 渲染动态页但重慢且需驱动；不应优先于接口，接口能打就别用。'},
          {'type': 'open', 'question': '你遇到一个页面 requests 抓空、Network 里也没找到数据接口。下一步你会怎么判断要不要上 Selenium？写 2-3 句。', 'answer': '先在 Network 仔细筛 XHR/Fetch 与 WebSocket，确认确实无接口；再考虑页面是否 iframe/加密；都无解才上 Selenium 渲染。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['SELENIUM', 'DYNAMIC', 'DRIVER']},
      ],
      tasks=[
          '找一个"查看源代码为空但页面有内容"的动态页，F12 确认它是否真无接口。',
          '（本机）装 selenium + ChromeDriver，写 5 行驱动浏览器打开一个页面并打印标题。',
          '把"动态页决策顺序"(先看接口→再 Selenium)记成你爬虫的默认 SOP。',
      ]),
])

# ============================ 第8章 反爬与合规 ============================
ch8 = CH('反爬与合规', [
    L('r8l1', 'robots.txt：先问人家让不让', '📄',
      r"""## 爬之前，先读"门牌告示"

每个正规站点根目录都有 `robots.txt`，声明**哪些路径允许爬、哪些禁止**。这是站长给爬虫的"规矩"。

```
# 例：https://x.com/robots.txt
User-agent: *
Disallow: /admin/        # 禁止爬后台
Disallow: /user/private/ # 禁止爬用户隐私区
Allow: /public/
```

### 怎么用
- 爬之前**先请求 `/robots.txt`** 读一遍
- 你目标路径在 `Disallow` 里 → 别爬，换数据源或放弃
- 有些站点直接 `Disallow: /` 表示"别爬任何东西"

### 态度
robots.txt 是**君子协定**，技术上你"能"无视它，但**尊重它是基本教养**，也是很多平台判断"善意/恶意"的依据。配合法律红线（下一节），别给自己找麻烦。""",
      figures=[{'key': 'robots_txt', 'caption': '📄 robots.txt 是网站根的"门牌告示"：Disallow 的路径就是人家说"别爬这里"，爬前先读'}],
      takeaway=r"""爬前先读 `/robots.txt`：Disallow 的路径别碰，Allow 的才爬。它是**君子协定**，技术上能无视但尊重它是基本教养，也是判断善恶爬虫的依据。""",
      words=[
          {'en': 'ROBOTS', 'zh': 'robots.txt：站点根的爬取规则告示'},
          {'en': 'DISALLOW', 'zh': 'Disallow：声明禁止爬取的路径'},
          {'en': 'ALLOW', 'zh': 'Allow：在禁止大前提下放开某些路径'},
      ],
      exercises=[
          {'type': 'choice', 'question': 'robots.txt 通常放在？', 'options': ['网站根目录 /robots.txt', '网页底部', '服务器密码文件'], 'answer': 0, 'explain': '标准约定放在站点根目录。'},
          {'type': 'choice', 'question': '目标路径出现在 `Disallow` 里，正确做法是？', 'options': ['别爬，换源或放弃', '偷偷爬', '无所谓'], 'answer': 0, 'explain': 'Disallow 即站长说禁止，应尊重。'},
          {'type': 'fill', 'question': 'robots.txt 里声明"禁止爬取"的指令关键字是 ______。', 'answer': 'disallow', 'explain': 'Disallow 列出禁止路径（大小写不敏感，答 disallow 即可）。'},
          {'type': 'tap', 'question': '关于 robots.txt，说法对的有？（多选）', 'options': ['声明允许/禁止路径', '爬前应先读', '是君子协定', '技术上无法无视'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'robots 声明规则、爬前该读、是君子协定；技术上可无视只是不道德。'},
          {'type': 'open', 'question': '有人说"robots.txt 又不是法律，我不看也行"。你怎么看？写 2-3 句。', 'answer': 'robots 虽非法定义务，但是行业基本教养与善意信号；无视它易被判定恶意、触发封禁甚至在法律纠纷中处于不利；尊重成本极低。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['ROBOTS', 'DISALLOW', 'ALLOW']},
      ],
      tasks=[
          '打开你目标站点的 /robots.txt，逐行读，标出和你目标数据相关的 Disallow 规则。',
          '写个小函数：请求 robots.txt 并简单判断某路径是否被禁止。',
          '如果你的目标在 Disallow 里，列出替代数据源或决定放弃的理由。',
      ]),

    L('r8l2', '速率限制与礼貌：做个有教养的爬虫', '🤝',
      r"""## 爬虫的"教养"清单

技术能爬，不等于该猛爬。有教养的爬虫长这样：

### 1. 控制频率
- 加随机延迟（第 7 章），别一秒几十次
- 控制并发数，别开几百个线程同时砸

### 2. 表明身份
- 用真实 UA，或在 UA 里留联系方式（有些站点欢迎"写明身份"的爬虫）
- 别伪装成普通用户骗人

### 3. 只取需要的
- 别把整个站镜像下载（那叫"搬站"，不是爬取）
- 缓存已抓的，重复请求用本地副本

### 4. 遵守 robots 与声明
- 第 8 章第 1 节讲过了，爬前读

### 一句话原则
> **只拿你要的、慢慢地拿、亮明身份地拿、人家不让的就不拿。** 这样你基本不会被人嫌弃，也能长期稳定地拿。""",
      takeaway=r"""有教养的爬虫四件套：**控频率(随机延迟/限并发)、亮身份(真实UA/留联系方式)、只取所需(不搬站/用缓存)、守 robots**。原则——只拿你要的、慢慢拿、亮明身份拿、不让的不拿。""",
      words=[
          {'en': 'POLITE', 'zh': '礼貌爬虫：控频、亮身份、只取所需、守规则'},
          {'en': 'CACHE', 'zh': '缓存：已抓的存本地，避免重复请求'},
          {'en': 'CONCURRENCY', 'zh': '并发：同时进行的请求数，过高像攻击'},
      ],
      exercises=[
          {'type': 'choice', 'question': '"搬站"（把整个站镜像下载）属于？', 'options': ['过度抓取，不提倡', '正常爬取', '必须做的'], 'answer': 0, 'explain': '无差别全站下载是滥用带宽，常被拒并封禁。'},
          {'type': 'choice', 'question': '在 UA 里留联系方式，主要作用是？', 'options': ['让站点方必要时能联系你、显得善意', '加快速度', '隐藏身份'], 'answer': 0, 'explain': '留联系方式便于站点方沟通，是善意信号。'},
          {'type': 'fill', 'question': '对已抓取过的内容存本地、避免重复请求，这叫______（填"缓存"或"加速"）。', 'answer': '缓存', 'explain': '缓存减少重复请求，既礼貌又高效。'},
          {'type': 'tap', 'question': '关于爬虫礼貌，做法对的有？（多选）', 'options': ['加随机延迟控频', '用真实 UA 亮身份', '只取需要的', '开几百线程猛砸'], 'multi': True, 'answer': [0, 1, 2], 'explain': '控频、亮身份、只取所需都礼貌；高并发猛砸是攻击行为。'},
          {'type': 'open', 'question': '你写了个爬虫每分钟打目标站 200 次，朋友说你"像在攻击"。你怎么改才能算礼貌？写 2-3 句。', 'answer': '降到合理频率（如每 1-3 秒一次）、限制并发、加随机延迟、必要时留 UA 联系方式，并确认未违反 robots。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['POLITE', 'CACHE', 'CONCURRENCY']},
      ],
      tasks=[
          '给爬虫加"频率上限"配置（如每分钟最多 N 次），并写进代码。',
          '在 UA 里加上你的标识/联系方式（如 "my-crawler/1.0 (contact: xx)"）。',
          '为你的爬虫加一层"已抓 URL 缓存"，避免重复请求同一页。',
      ]),

    L('r8l3', '法律红线与隐私：哪些绝对不能碰', '🚫',
      r"""## 这一节，是来泼冷水的

技术玩得再溜，踩了线就是另一个故事。几条**硬红线**：

### ⛔ 绝对别碰
- **个人信息**：身份证、手机号、住址、行踪、聊天记录——受《个人信息保护法》严管，未经同意抓取/贩卖可能刑事追责
- **付费 / 登录墙内的专有内容**：别人卖钱的课程、会员资源，你爬了等于盗窃
- **国家秘密、色情、赌博、暴恐等违法内容**：碰了就不是封号，是办案
- **突破技术措施**：人家明显设防（加密、收费、验证码防爬），你专门去"破解"绕过，法律风险陡增

### ⚠️ 要小心
- **版权**：抓来的文章/图片/视频，别擅自商用转发
- **不正当竞争**：用爬虫把对手数据整碗端走做同种生意，可能构成不正当竞争

### 自我三问（动手前默念）
1. 这数据**公开且无个人敏感信息**吗？
2. 人家 robots / 服务条款**允许**吗？
3. 我拿去**怎么用**，会不会损害别人？

三问过了，再爬。过不了，收手。""",
      figures=[{'key': 'legal_redline', 'caption': '🚫 红线：个人信息/付费内容/违法内容/破防绕过 千万别碰；版权与不正当竞争也要小心'}],
      takeaway=r"""硬红线：个人信息、付费专有内容、违法内容、专门破解技术措施——碰了可能刑事追责。版权/不正当竞争也要小心。**动手前三问：公开无敏感？允许？用途正当？** 过不了就收手。""",
      words=[
          {'en': 'PIPL', 'zh': '个人信息保护法：严管个人数据的收集与使用'},
          {'en': 'COPYRIGHT', 'zh': '版权：抓来的内容别擅自商用转发'},
          {'en': 'RED LINE', 'zh': '红线：违法/侵权/破防行为，碰了后果严重'},
      ],
      exercises=[
          {'type': 'tap', 'question': '以下哪些属于"绝对别碰"的硬红线？（多选）', 'options': ['抓取他人手机号/身份证', '爬取付费会员专有内容', '爬取公开的新闻标题', '专门破解对方技术防爬措施'], 'multi': True, 'answer': [0, 1, 3], 'explain': '个人信息、付费专有内容、破解技术措施都是红线；公开新闻一般无碍（仍看版权/robots）。'},
          {'type': 'choice', 'question': '动手爬之前，自我三问不包含以下哪条？', 'options': ['我写得快不快', '数据公开且无敏感信息吗', '人家允许吗', '用途正当吗'], 'answer': 0, 'explain': '三问是公开无敏感、允许、用途正当；"写得快不快"无关合规。'},
          {'type': 'fill', 'question': '严管个人数据、违规可追责的法律叫《个人信息______法》。', 'answer': '保护', 'explain': '《个人信息保护法》(PIPL) 规范个人数据处理。'},
          {'type': 'tap', 'question': '关于版权与不正当竞争，说法对的有？（多选）', 'options': ['抓来的内容别擅自商用', '整碗端走对手数据做同种生意有风险', '公开数据就能随便用', '应注意服务条款'], 'multi': True, 'answer': [0, 1, 3], 'explain': '商用转发涉版权、端走对手数据涉不正当竞争、应注意条款；"公开就能随便用"是误区。'},
          {'type': 'open', 'question': '你想爬一个论坛的全部用户发帖做"舆情分析"并打算公开发报告。至少哪两点让你必须谨慎？写 2-3 句。', 'answer': '发帖可能含个人信息需脱敏、公开报告涉及隐私与版权；须获授权/匿名化并遵守平台条款，否则违法风险高。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['PIPL', 'COPYRIGHT', 'RED LINE']},
      ],
      tasks=[
          '用"自我三问"逐条评估你的目标数据：公开无敏感？允许？用途正当？写下结论。',
          '查一下你目标站点的"服务条款/使用协议"，看是否禁止爬取。',
          '列出你这个项目"绝对不抓"的三类数据，作为个人红线清单。',
      ]),
])

# ============================ 第9章 实战项目 ============================
ch9 = CH('实战项目', [
    L('r9l1', '项目一：造一个天气看板', '🌤️',
      r"""## 目标：一句话看遍明天天气

这节咱们把前面学的**打接口 + 取字段 + 存文件**串成一个真东西：一个天气看板。

### 真实做法（拿 wttr.in 当免费接口，无需 key）
```python
import requests
city = 'Chengdu'
url = f'https://wttr.in/{city}?format=j1'   # 返回 JSON
r = requests.get(url, timeout=10)
data = r.json()
cur = data['current_condition'][0]
print(cur['temp_C'], '℃', cur['lang_zh'][0]['value'])
```
- `?format=j1` 是 wttr.in 的 JSON 模式
- `r.json()` 把响应正文直接转成字典/列表
- 再用 `data['current_condition'][0]` 一层层往下取

### 升级成"看板"
把多次查询的结果攒进一个字典，最后 `json.dump` 存成 `weather.json`，配合定时任务（第 10 章）就能每天自动更新。

> 下面代码框是**离线模拟版**：用一个本地 dict 假装是接口返回的 JSON，练的就是"取字段 + 存文件"的手感。""",
      figures=[{'key': 'api_json', 'caption': '🌤️ 天气看板链路：构造带参 URL → requests 拿 JSON → 取 temp_C/天气描述 → 攒进字典 → dump 成 weather.json'}],
      takeaway=r"""天气看板 = 构造带参 URL + requests.get + r.json() 取字段 + 攒字典 + json.dump 存盘。wttr.in 的 ?format=j1 免 key 直接给 JSON，是练接口爬取的神仙站点。""",
      words=[
          {'en': 'ENDPOINT', 'zh': '接口地址：API 的 URL，发请求拿数据的地方'},
          {'en': 'JSON DUMP', 'zh': 'json.dump：把 Python 数据写成 JSON 文件'},
          {'en': 'TIMEOUT', 'zh': '超时：requests 等响应的秒数，防卡死'},
      ],
      code="""# 离线模拟：假设这就是接口返回的 JSON
fake_api = {
    'current_condition': [{'temp_C': '28', 'lang_zh': [{'value': '晴'}]}],
    'city': 'Chengdu',
}
cur = fake_api['current_condition'][0]
result = {
    'city': fake_api['city'],
    'temp': cur['temp_C'],
    'desc': cur['lang_zh'][0]['value'],
}
print('天气:', result['city'], result['temp'] + '℃', result['desc'])
# 真实项目用 json.dump(result, open('weather.json','w',encoding='utf-8')) 存盘；这里直接看结果字典
print('存盘内容:', result)""",
      exercises=[
          {'type': 'choice', 'question': 'wttr.in 要拿到 JSON 格式，URL 里要加？', 'options': ['?format=j1', '?mode=html', '?type=xml'], 'answer': 0, 'explain': 'wttr.in 的 ?format=j1 专门返回 JSON。'},
          {'type': 'choice', 'question': 'requests 把响应正文转成 Python 字典，用？', 'options': ['r.json()', 'r.text', 'r.content'], 'answer': 0, 'explain': 'r.json() 自动解析 JSON；r.text 是字符串，r.content 是字节。'},
          {'type': 'fill', 'question': '给 requests 加超时、最多等 10 秒，应写 requests.get(url, ______=10)。', 'answer': 'timeout', 'explain': 'timeout 防止请求卡死一直等。'},
          {'type': 'fill', 'question': '把字典存成 JSON 文件，用 json.______(data, f)（填 dump/load）。', 'answer': 'dump', 'explain': 'json.dump 写盘；json.load 才是读。'},
          {'type': 'order', 'question': '天气看板的数据链路排排序：', 'steps': ['构造带参 URL', 'requests.get 拿响应', 'r.json() 转字典', '取字段攒结果', 'json.dump 存盘'], 'explain': '顺序是：构造 URL → 发请求 → 转 JSON → 取字段 → 存盘。'},
          {'type': 'tap', 'question': '关于这个天气项目，说法对的有？（多选）', 'options': ['用 r.json() 解析响应', '可用 json.dump 存盘', '需要自己猜编码', '可配定时任务每天更新'], 'multi': True, 'answer': [0, 1, 3], 'explain': 'r.json 解析、json.dump 存盘、定时任务都正确；JSON 接口一般已是 utf-8，不必手猜编码。'},
          {'type': 'coding', 'question': '下面 fake_api 里取出 desc 并打印，目标输出包含"晴"。', 'starter': "fake_api = {'current_condition': [{'lang_zh': [{'value': '晴'}]}]}\n# 取出 desc 并打印", 'expect': '晴', 'hint': 'cur = fake_api["current_condition"][0]; print(cur["lang_zh"][0]["value"])'},
          {'type': 'open', 'question': '如果要同时查 3 个城市并存一张表，你的数据结构和循环思路怎么设计？写 3-4 句。', 'answer': '用一个列表 cities 循环，每次请求一个城市、取出 temp/desc 塞进一个字典，再把所有字典 append 到一个列表，最后 json.dump 整个列表成 weather.json。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['ENDPOINT', 'TIMEOUT', 'JSON']},
      ],
      tasks=[
          '用 wttr.in 真实跑一次，把成都/北京/上海的当前温度抓出来打印。',
          '把多城市结果攒进一个列表，json.dump 存成 weather.json。',
          '试着加一个"未来 3 天"的字段提取（wttr.in 的 weather 数组里）。',
      ]),

    L('r9l2', '项目二：把榜单抓成一张表', '📊',
      r"""## 目标：书单 / 电影榜 / 商品榜 → 一键导出 CSV

很多网站有"排行榜"，咱们用 BeautifulSoup 把每条记录的字段抠出来，存成 Excel 能直接打开的 CSV。

### 真实做法
```python
import requests
from bs4 import BeautifulSoup
import csv

html = requests.get(url, timeout=10).text
soup = BeautifulSoup(html, 'lxml')
rows = []
for item in soup.select('.rank-item'):
    title = item.select_one('.title').get_text(strip=True)
    score = item.select_one('.score').get_text(strip=True)
    rows.append({'title': title, 'score': score})

with open('rank.csv', 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=['title', 'score'])
    w.writeheader()
    w.writerows(rows)
```
- `encoding='utf-8-sig'`：让 Excel 打开中文不乱码的关键（带 BOM）
- `newline=''`：避免 Windows 下出现空行

> 下面代码框是**离线模拟版**：用一段本地 HTML 字符串当"被抓的页面"，练的就是"按选择器抠字段 + 写 CSV"的手感。""",
      figures=[{'key': 'data_store', 'caption': '📊 榜单→表：requests 拿 HTML → bs4 按 .rank-item 抠 title/score → DictWriter 写 csv(utf-8-sig 防 Excel 乱码)'}],
      takeaway=r"""榜单抓取 = requests 拿 HTML + bs4.select 抠每条 + DictWriter 写 csv（务必 utf-8-sig + newline='' 让 Excel 友好）。把"网页列表"变"可分析表格"就是这么朴实。""",
      words=[
          {'en': 'SELECT', 'zh': 'select：按 CSS 选择器批量找元素'},
          {'en': 'DICTWRITER', 'zh': 'DictWriter：把字典列表写成 CSV 的列'},
          {'en': 'BOM', 'zh': 'utf-8-sig：带 BOM 的 utf-8，Excel 打开中文不乱码'},
      ],
      code="""# 离线模拟：用一段本地 HTML 字符串当"被抓的页面"
# 真实项目用 bs4 的 select/select_one 抠字段；这里用纯字符串切片演示同一思路
html = '<ul><li class="rank-item"><span class="title">黑客与画家</span><span class="score">9.1</span></li><li class="rank-item"><span class="title">代码大全</span><span class="score">9.3</span></li></ul>'

rows = []
for chunk in html.split('<li class="rank-item">'):
    if '</span>' not in chunk:
        continue
    t_s = chunk.find('<span class="title">') + len('<span class="title">')
    t_e = chunk.find('</span>', t_s)
    s_s = chunk.find('<span class="score">') + len('<span class="score">')
    s_e = chunk.find('</span>', s_s)
    rows.append((chunk[t_s:t_e].strip(), chunk[s_s:s_e].strip()))

print('抓到', len(rows), '条')
for title, score in rows:
    print(title, score)

# 模拟写 CSV（真实项目用 open('rank.csv','w',encoding='utf-8-sig',newline='') + csv）
lines = ['title,score']
for title, score in rows:
    lines.append(title + ',' + score)
print('--- CSV ---')
for line in lines:
    print(line)""",
      exercises=[
          {'type': 'choice', 'question': '让 Excel 打开中文 CSV 不乱码，编码应写？', 'options': ['utf-8-sig', 'ascii', 'gb2312'], 'answer': 0, 'explain': 'utf-8-sig 带 BOM，Excel 识别为 utf-8。'},
          {'type': 'choice', 'question': '`soup.select(".rank-item")` 返回的是？', 'options': ['所有匹配的元素列表', '第一个匹配', '属性值'], 'answer': 0, 'explain': 'select 返回全部匹配元素的列表；select_one 才返回第一个。'},
          {'type': 'fill', 'question': '按 CSS 选择器找"第一个"元素，用 soup._______one(sel)（填 select）。', 'answer': 'select', 'explain': 'select_one 取第一个匹配元素。'},
          {'type': 'fill', 'question': '取元素里的文字并去掉首尾空格，用 el.get_text(______=True)（填 strip）。', 'answer': 'strip', 'explain': 'get_text(strip=True) 返回去空白的文本。'},
          {'type': 'order', 'question': '榜单抓成表的步骤排排序：', 'steps': ['requests 拿 HTML', 'BeautifulSoup 解析', 'select 抠每条字段', 'DictWriter 写 csv'], 'explain': '拿 HTML → 解析 → 抠字段 → 写 CSV，顺序别乱。'},
          {'type': 'tap', 'question': '关于 CSV 导出，说法对的有？（多选）', 'options': ['用 utf-8-sig 防 Excel 乱码', 'newline="" 防空行', '用 csv.DictWriter', '用 print 直接写文件'], 'multi': True, 'answer': [0, 1, 2], 'explain': '前三项都是正确姿势；直接 print 写文件不带表头且格式乱。'},
          {'type': 'coding', 'question': '下面用纯字符串切片，从 html 里取出 class=title 的文字并打印，目标输出含"黑客与画家"。', 'starter': "html = '<p class=\"title\">黑客与画家</p>'\n# 用字符串切片取出 > 和 </ 之间的文字并打印", 'expect': '黑客与画家', 'hint': "s = html.find('>') + 1; e = html.find('</'); print(html[s:e])"},
          {'type': 'open', 'question': '如果某条榜单缺了 score 字段（网站偶发），你的代码怎么不崩？写 2-3 句。', 'answer': '用 select_one 后判断是否为 None，或用 try/except 兜底给默认值 0/空串，保证循环不中断。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['SELECT', 'DICTWRITER', 'BOM']},
      ],
      tasks=[
          '找一个你喜欢的榜单页（电影/图书/游戏），用浏览器开发者工具确认它的列表选择器。',
          '用 bs4 + csv 把该榜单前 20 条导出成 rank.csv，用 Excel 打开确认不乱码。',
          '加一列"抓取时间"，让表格自带时间戳。',
      ]),

    L('r9l3', '项目三：批量下载一整套图片', '🖼️',
      r"""## 目标：把图集 / 表情包一次性搬回家

图片下载和"抓文字"只差一步：**响应是二进制，要用 `wb` 模式写文件**。

### 真实做法
```python
import requests, os
os.makedirs('imgs', exist_ok=True)
for i, src in enumerate(img_urls):
    r = requests.get(src, timeout=10)
    with open(f'imgs/{i}.jpg', 'wb') as f:
        f.write(r.content)     # 注意是 content（字节），不是 text
```

### 关键两点
- `r.content`：二进制正文（图片本体）；`r.text` 是给文字用的，图片必须用 `content`
- `wb`：以"写二进制"方式开文件，否则图片会坏

> 取图片链接这一步，还是用 bs4 抠 `<img src>`（上一节练过）。下面代码框模拟"拿到链接后怎么存"，离线可跑。""",
      figures=[{'key': 'download_flow', 'caption': '🖼️ 批量下图片：bs4 抠 img src → 循环 requests.get → r.content 二进制 → open(...,"wb") 写盘'}],
      takeaway=r"""批量下图片 = bs4 抠 img src + 循环 requests.get + 用 r.content（字节）配合 open(...,'wb') 写盘。文字用 text、图片用 content，这一步搞反图片就废了。""",
      words=[
          {'en': 'CONTENT', 'zh': 'r.content：响应的二进制正文，用于图片/文件'},
          {'en': 'WB MODE', 'zh': 'wb：以写二进制方式开文件，存图片必须用'},
          {'en': 'Makedirs', 'zh': 'os.makedirs：递归建目录，exist_ok 防报错'},
      ],
      code="""# 离线模拟：假设已经用 bs4 抠到了这些图片链接（真实情况从网页提取）
img_urls = ['https://x.com/a.jpg', 'https://x.com/b.jpg', 'https://x.com/c.jpg']
# 真实项目：先 os.makedirs('imgs', exist_ok=True)，再循环
#   r = requests.get(src, timeout=10); open(f'imgs/{i}.jpg','wb').write(r.content)
for i, src in enumerate(img_urls, 1):
    print(f'下载第 {i} 张 -> imgs/{i}.jpg  (来自 {src})')
print('共下载', len(img_urls), '张')""",
      exercises=[
          {'type': 'choice', 'question': '下载图片该用响应哪个属性？', 'options': ['r.content（字节）', 'r.text（字符串）', 'r.url'], 'answer': 0, 'explain': '图片是二进制，必须用 r.content；r.text 会变成乱码字符串。'},
          {'type': 'choice', 'question': '写图片文件应以哪种模式打开？', 'options': ['wb（写二进制）', 'w（写文本）', 'r（读）'], 'answer': 0, 'explain': '图片是二进制，wb 才不会损坏文件。'},
          {'type': 'fill', 'question': '递归创建目录且已存在不报错，用 os.______("imgs", exist_ok=True)（填 makedirs）。', 'answer': 'makedirs', 'explain': 'os.makedirs 建目录，exist_ok=True 避免重复创建报错。'},
          {'type': 'fill', 'question': '批量下载时给文件编号，通常用 for i, src in enumerate(______)。', 'answer': 'img_urls', 'explain': 'enumerate 同时拿到下标和内容，方便编号命名。'},
          {'type': 'tap', 'question': '关于图片批量下载，做法对的有？（多选）', 'options': ['用 r.content 取字节', 'open 用 wb 模式', '循环每个链接分别下载', '用 r.text 存图片'], 'multi': True, 'answer': [0, 1, 2], 'explain': 'content+wb+循环都正确；r.text 存图片会损坏。'},
          {'type': 'order', 'question': '批量下图片的流程排排序：', 'steps': ['bs4 抠 img src', '循环每个链接', 'requests.get 拿 content', 'open wb 写盘'], 'explain': '抠链接 → 循环 → 取 content → 写盘。'},
          {'type': 'coding', 'question': '下面把 img_urls 里每个链接打印成"第N张:链接"格式，目标输出含"第1张"。', 'starter': "img_urls = ['a.jpg', 'b.jpg']\n# 打印 第1张:a.jpg / 第2张:b.jpg", 'expect': '第1张', 'hint': 'for i, src in enumerate(img_urls, 1): print(f"第{i}张:{src}")'},
          {'type': 'open', 'question': '批量下载几百张图，怎么避免把人家服务器打爆、也避免自己被封？写 2-3 句。', 'answer': '加随机延迟与并发上限、设超时与重试、尊重 robots、只下需要的分辨率，必要时限速。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['CONTENT', 'WB MODE', 'Makedirs']},
      ],
      tasks=[
          '找一个图集页，用 bs4 抠出全部 `<img>` 的 src（注意相对路径要拼成绝对 URL）。',
          '写循环把前 10 张下到 imgs/ 目录，确认图片能正常打开。',
          '加上随机延迟（如 time.sleep(random.uniform(0.5,1.5))）避免高频请求。',
      ]),

    L('r9l4', '综合毕业项目：挑一个站，从头爬到尾', '🚀',
      r"""## 现在，轮到你了

前面三个项目各练了一块。毕业项目把你学到的**全部**串起来：选一个你真感兴趣的公开站点，从"定目标 → 看数据源 → 抓 → 解析 → 存 → 守规矩"走完整一遍。

### 毕业项目 Checklist
1. **定目标**：明确要抓什么字段（标题/价格/链接/图片…），要存成什么（CSV/JSON/Excel）
2. **看数据源**：F12 → Network，确认数据是写在 HTML 还是接口 JSON
3. **写抓取**：requests + 必要的 headers/参数
4. **写解析**：bs4 或正则或 r.json()，把字段抠出来
5. **写存储**：csv / json / openpyxl，带好编码
6. **加礼貌**：随机延迟、真实 UA、遵守 robots
7. **做兜底**：try/except 接住脏数据，缺字段给默认
8. **写说明**：README 写清"这个脚本干啥、怎么跑、数据用途"

### 选题建议（由易到难）
- 易：静态博客的文章标题+日期列表
- 中：电商/图书榜单（bs4 抠字段 + CSV）
- 难：需要登录态或分页翻页的新闻聚合

> 卡住很正常。把报错整段贴给小光（或搜索引擎），90% 的坑前人踩过。这一节没有标准答案，完成你自己那一个就行。""",
      figures=[{'key': 'project_map', 'caption': '🚀 毕业项目全链路：定目标→看数据源→抓→解析→存→加礼貌→做兜底→写说明，八步走完即毕业'}],
      takeaway=r"""毕业项目 = 八步闭环：定目标→看数据源→抓→解析→存→加礼貌→做兜底→写说明。没有标准答案，选一个你真感兴趣的公开站，从头爬到尾就是毕业。""",
      words=[
          {'en': 'PIPELINE', 'zh': '流水线：抓取→解析→存储一气呵成的完整链路'},
          {'en': 'README', 'zh': '说明文档：写清脚本用途、运行方式、数据用途'},
          {'en': 'ROBUST', 'zh': '健壮性：try/except + 默认值，脏数据不崩'},
      ],
      exercises=[
          {'type': 'choice', 'question': '毕业项目第一步应该是？', 'options': ['明确要抓什么字段、存成什么', '直接写代码', '先买服务器'], 'answer': 0, 'explain': '先定目标（字段+存储格式）再动手，否则容易返工。'},
          {'type': 'tap', 'question': '毕业 Checklist 里"守规矩"相关的有？（多选）', 'options': ['随机延迟/真实 UA', '遵守 robots', 'try/except 兜底', '写 README 说明用途'], 'multi': True, 'answer': [0, 1, 3], 'explain': '延迟/UA/robots/README 都属"守规矩与可维护"；try/except 是健壮性。'},
          {'type': 'fill', 'question': '判断数据是写在 HTML 还是接口 JSON，要用开发者工具的 ______ 面板。', 'answer': 'network', 'explain': 'Network 面板能看到每个请求的响应内容，据此判断数据源。'},
          {'type': 'open', 'question': '说说你打算做的毕业项目：选哪个站、抓什么字段、存成什么格式、怎么保证不踩红线？写 4-6 句。', 'answer': '示例：选某公开图书榜，抓书名/作者/评分存 CSV；数据源是 HTML 用 bs4；加随机延迟与真实 UA、查过 robots 允许；不抓个人信息、不商用。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['PIPELINE', 'README', 'ROBUST']},
      ],
      tasks=[
          '在纸上（或笔记）写出你的毕业项目八步计划，每一步一句话。',
          '用 F12 把目标站的数据源确认清楚（HTML 还是接口）。',
          '写出可运行的 v1 脚本，至少抓到 10 条数据并成功存盘。',
          '给脚本加 try/except 兜底和随机延迟，跑通一整轮。',
      ]),
])

# ============================ 第10章 毕业冲刺 ============================
ch10 = CH('毕业冲刺', [
    L('r10l1', '调试与排错：爬虫最常见的坑', '🐞',
      r"""## 写完跑不起来？这节专治各种"为什么没数据"

爬虫 90% 的时间在**调试**。提前认识这几个坑，能省你一整周。

### 坑 1：拿到空列表
- 原因：选择器写错 / 网站改版 / 数据是 JS 动态加载（HTML 里根本没有，要去 Network 找接口）
- 排查：`print(soup)` 看原始 HTML，确认目标在不在里面

### 坑 2：403 Forbidden
- 原因：没带 UA，被当成机器人；或触发了反爬
- 排查：加上 `headers={'User-Agent': '...'}`；还不行看是不是要带 Cookie/Session（第 7 章）

### 坑 3：中文乱码
- 原因：编码猜错
- 排查：手动 `r.encoding='utf-8'`（或 gbk），再 `r.text`

### 坑 4：字段缺失崩了
- 原因：某条记录没有这个字段
- 排查：`try/except` 或 `el.get_text() or ''` 兜底

### 坑 5：爬着爬着被封
- 原因：太快/太猛/不守 robots
- 排查：降速、加代理轮换、装"礼貌爬虫"

### 通用排错套路
> **缩小范围**：先打印原始响应 → 再打印解析中间结果 → 哪一步变空，坑就在哪一步。别一上来就怀疑人生。""",
      figures=[{'key': 'debug_wheel', 'caption': '🐞 排错套路：打印原始响应→打印解析中间结果→定位变空的那一步；常见坑=空列表/403/乱码/缺字段/被封'}],
      takeaway=r"""爬虫排错套路：先 print 原始响应，再 print 解析中间结果，定位"哪一步变空"。五大坑：空列表(选择器/动态加载)、403(补UA)、乱码(设encoding)、缺字段(try兜底)、被封(降速装礼貌)。""",
      words=[
          {'en': 'DEBUG', 'zh': '调试：逐层打印、缩小范围定位问题'},
          {'en': 'DYNAMIC', 'zh': '动态加载：数据由 JS 现拉，HTML 里没有，要去接口拿'},
          {'en': 'FALLBACK', 'zh': '兜底：try/except 或默认值，脏数据不崩'},
      ],
      code="""# 离线模拟：用一组"记录"（有的故意缺 price）演示"字段缺失兜底"
items = [
    {'name': 'A', 'price': '39'},
    {'name': 'B'},            # 故意缺 price
    {'name': 'C', 'price': '25'},
]
# 套路：先确认原始里有没有目标
print('总记录数:', len(items))
for it in items:
    # 兜底：没有 price 也不崩
    name = it.get('name', '无名')
    price = it.get('price', '暂无')
    print(name, price)""",
      exercises=[
          {'type': 'choice', 'question': '抓到空列表，第一步该？', 'options': ['print 原始 HTML 看目标在不在', '直接重写整个脚本', '放弃'], 'answer': 0, 'explain': '先确认原始响应里有没有目标，再判断是选择器问题还是动态加载。'},
          {'type': 'choice', 'question': '遇到 403 Forbidden，最可能先试？', 'options': ['加 User-Agent 请求头', '换电脑', '多刷几次'], 'answer': 0, 'explain': '403 常因缺 UA 被当机器人，补 headers 是最常见解法。'},
          {'type': 'fill', 'question': '中文乱码时，手动指定编码写 r.______ = "utf-8"（填 encoding）。', 'answer': 'encoding', 'explain': '设 r.encoding 让 requests 用正确编码解码。'},
          {'type': 'tap', 'question': '关于"动态加载"，说法对的有？（多选）', 'options': ['数据由 JS 现拉，HTML 里没有', '要去 Network 找接口', 'bs4 一定能解析', '可能要打 JSON 接口'], 'multi': True, 'answer': [0, 1, 3], 'explain': '动态加载数据不在初始 HTML，要去接口拿；bs4 解析 HTML 救不了动态加载。'},
          {'type': 'fill', 'question': '字段可能缺失，用 try/______ 兜底避免崩溃（填 except）。', 'answer': 'except', 'explain': 'try/except 接住可能缺失的字段访问。'},
          {'type': 'order', 'question': '通用排错套路排排序：', 'steps': ['打印原始响应', '打印解析中间结果', '定位变空的那一步', '针对性修复'], 'explain': '逐层打印缩小范围：原始→中间结果→定位空步→修复。'},
          {'type': 'open', 'question': '你的脚本抓回来 100 条里只有 3 条有数据，其余是 None。你按什么顺序排查？写 3-4 句。', 'answer': '先打印几条原始 HTML 确认目标结构；再看选择器是否只在部分条目生效；最后查是不是部分条目缺字段需要兜底。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['DEBUG', 'DYNAMIC', 'FALLBACK']},
      ],
      tasks=[
          '故意把某个选择器写错，观察"空列表"现象，再用 print 原始 HTML 定位。',
          '给一个 requests 请求加 User-Agent，对比加之前是否还 403。',
          '写一段"字段缺失兜底"代码，让缺 price 的条目输出"暂无"而非崩溃。',
      ]),

    L('r10l2', '毕业地图与下一步', '🗺️',
      r"""## 你已经能写爬虫小脚本了

回头看这条路线：Python 基本功 → 看懂网页/HTTP → requests 抓取 → BeautifulSoup 解析 → 正则/JSON 兜底 → 存储 → 实战项目 → 守规矩。**完整闭环，你已经走完。**

### 知识地图
- ✅ 发请求：GET/POST、headers、参数、超时、Session
- ✅ 解析：bs4 选择器、正则、r.json()
- ✅ 存储：CSV / JSON / Excel
- ✅ 工程化：随机延迟、代理、重试、异常处理
- ✅ 合规：robots、礼貌、法律红线

### 下一步可以往哪走
- **Scrapy**：专业爬虫框架，自带调度/管道，适合大批量
- **异步爬虫**：`asyncio` + `aiohttp`，并发量量级提升
- **爬动态页**：`Selenium` / `Playwright` 驱动真实浏览器，专治 JS 渲染
- **数据可视化**：`pandas` + `matplotlib` 把抓来的数据画成图
- **定时任务**：把脚本挂到 cron / 云函数，每天自动跑

### 给毕业的你一句话
> 爬虫不是终点，是"把网上信息变成你能用的数据"的第一把钥匙。带着**尊重规则 + 解决实际问题**的心态去用，它会是你最趁手的工具之一。去写你的第一个真项目吧 🎉""",
      figures=[{'key': 'roadmap', 'caption': '🗺️ 毕业地图：请求→解析→存储→工程化→合规 全通；下一步 Scrapy/异步/Selenium/可视化/定时任务'}],
      takeaway=r"""你已走完完整闭环：请求→解析→存储→工程化→合规。下一步可探 Scrapy、异步 aiohttp、Selenium 爬动态页、pandas 可视化、定时任务。带着"守规则+解决实际问题"去写你的第一个真项目。""",
      words=[
          {'en': 'SCRAPY', 'zh': 'Scrapy：专业爬虫框架，自带调度与管道'},
          {'en': 'ASYNC', 'zh': '异步：asyncio+aiohttp，并发量级提升'},
          {'en': 'SELENIUM', 'zh': 'Selenium：驱动真实浏览器，爬 JS 动态页'},
      ],
      exercises=[
          {'type': 'choice', 'question': '爬"JS 动态渲染、HTML 里没数据"的页面，该用？', 'options': ['Selenium/Playwright 驱动浏览器', '更长的 time.sleep', '换个网站'], 'answer': 0, 'explain': '动态页要等 JS 执行，需真实浏览器驱动。'},
          {'type': 'choice', 'question': '要并发量大幅提升，应学？', 'options': ['asyncio + aiohttp', '多写 print', '用 Excel'], 'answer': 0, 'explain': '异步 IO 能大幅提升并发请求效率。'},
          {'type': 'tap', 'question': '以下哪个是专业爬虫框架？（多选）', 'options': ['Scrapy', 'requests', 'Selenium', 'aiohttp'], 'multi': True, 'answer': [0, 2, 3], 'explain': 'Scrapy 框架、Selenium 浏览器驱动、aiohttp 异步都属进阶；requests 是入门库。'},
          {'type': 'fill', 'question': '把抓来的数据画成图，常用 ______ + matplotlib（填 pandas）。', 'answer': 'pandas', 'explain': 'pandas 做数据处理，matplotlib 画图。'},
          {'type': 'open', 'question': '学完这套课，你最想马上用爬虫解决自己的哪个具体问题？写 2-3 句。', 'answer': '开放题：例如自动汇总每周关注的几个站点更新、把喜欢的图集备份、做个人天气/汇率看板等。'},
          {'type': 'typing', 'question': '敲一敲今天的关键词：', 'words': ['SCRAPY', 'ASYNC', 'SELENIUM']},
      ],
      tasks=[
          '在 README 里写下你的"爬虫知识地图"自测（哪些已掌握、哪些待学）。',
          '挑一个进阶方向（Scrapy / 异步 / Selenium 之一），跑通官方最小示例。',
          '把你真正的第一个毕业项目脚本整理好，加上说明和合规声明，留作纪念。',
      ]),
])

# ===== 章节插入点（后续章节用 Edit 追加到这里之前） =====
CHAPTERS = [ch0, ch1, ch2, ch3, ch4, ch5, ch6, ch7, ch8, ch9, ch10]
# ===== 章节插入点结束 =====


def collect_words(chapters):
    seen = {}
    for ch in chapters:
        for les in ch['lessons']:
            for w in les.get('words') or []:
                seen[w['en'].upper()] = w
    return list(seen.values())


def write_all(chapters):
    # 自动补音标：对没有 pron 的单词用 eng_to_ipa 生成（保留手工音标）
    for ch in chapters:
        for les in ch['lessons']:
            for w in les.get('words') or []:
                if not w.get('pron') and w.get('en'):
                    p = _to_ipa(w['en'])
                    if p:
                        w['pron'] = p

    # 注入章节配色，保证同章节卡片风格统一
    for ch in chapters:
        c = CHAPTER_COLOR.get(ch['title'])
        if c:
            for les in ch['lessons']:
                if 'color' not in les:
                    les['color'] = c
    course = {'title': '小光陪你写爬虫（14+ 实战）', 'chapters': chapters}
    os.makedirs(DATA, exist_ok=True)

    with open(os.path.join(DATA, 'course.js'), 'w', encoding='utf-8') as f:
        f.write('// 自动生成，请勿手改。重跑 tools/build_course.py 即可更新。\n')
        f.write('window.COURSE_DATA = ')
        f.write(json.dumps(course, ensure_ascii=False, indent=1))
        f.write(';\n')

    with open(os.path.join(DATA, 'talk.js'), 'w', encoding='utf-8') as f:
        f.write('// 本套课无视频，讲一讲字幕留空（朗读讲义 / 学习成果仍可用浏览器语音）。\n')
        f.write('window.LESSON_TALK = {};\n')

    with open(os.path.join(DATA, 'audio.js'), 'w', encoding='utf-8') as f:
        f.write('// 未烘焙真人语音时留空，前端自动回退浏览器 TTS。\n')
        f.write('window.AUDIO_MAP = {};\n')

    words = collect_words(chapters)
    with open(os.path.join(DATA, 'words.js'), 'w', encoding='utf-8') as f:
        f.write('// 打字练习词库（由课程术语自动汇总）。\n')
        f.write('window.WORD_LIST = ')
        f.write(json.dumps(words, ensure_ascii=False, indent=1))
        f.write(';\n')

    with open(os.path.join(DATA, 'vocab-words.js'), 'w', encoding='utf-8') as f:
        f.write('// 本课单词直接写在 course.js 各 lesson.words，这里留空避免重复。\n')
        f.write('window.VOCAB_WORDS = {};\n')

    # 统计
    n_lessons = sum(len(ch['lessons']) for ch in chapters)
    n_ex = sum(len(les.get('exercises') or []) for ch in chapters for les in ch['lessons'])
    n_tasks = sum(len(les.get('tasks') or []) for ch in chapters for les in ch['lessons'])
    print('课程生成完成：%d 章 / %d 节 / %d 道题 / %d 个任务' % (len(chapters), n_lessons, n_ex, n_tasks))


if __name__ == '__main__':
    write_all(CHAPTERS)
